'''module to transfer excel file into sqlite db'''
import sqlite3
import openpyxl
from openpyxl.utils import get_column_letter

# connect DB
CONN = sqlite3.connect('nr.db')
CURSOR = CONN.cursor()


def load_data(file):
    '''load excel file'''
    # connect excel
    work_book = openpyxl.load_workbook(file)
    sheets = work_book.get_sheet_names()

    for sheet_name in sheets:

        sheet = work_book.get_sheet_by_name(sheet_name)

        # read titles
        max_col = sheet.max_column
        titles = [sheet[get_column_letter(i) + '1'].value
                  for i in range(1, max_col + 1)]

        # print(titles)

        # prepare insert_list
        insert_list = []
        max_row = sheet.max_row
        print(max_row)
        for j in range(2, max_row + 1):
            row = tuple([sheet[get_column_letter(i)
                               + str(j)].value for i in range(1, max_col + 1)])
            print(row)
            insert_list.append(row)

        # print(insert_list)
        # insert list into sql
        # prepare string
        question_mark = '(' + "?," * (len(titles) - 1) + '?)'
        string = "INSERT INTO {0} VALUES {1}".format(sheet_name, question_mark)
        # print(string)

        CURSOR.executemany(string, insert_list)

    # close connection to DB
    CONN.commit()
    CONN.close()


if __name__ == "__main__":
    load_data('00_Collector.xlsx')
