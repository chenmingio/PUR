from openpyxl import load_workbook

import os
from pur_doc import constant, sql, word, xls_inject

from pprint import pprint

# test projects
test_project_blank = ""
test_project_none = None
test_project_1 = "1111E.001236"
test_project_fake = "fake_project_id"

# test single parts

# project_vendor_tuple1
project_vendor_tuple1 = ("1111E.001169", "48200041")
project_vendor_tuple2 = ("1361P.000054", "49100612")
# test project/part tuple
test_project_part_tuple_1 = ("1111E.001236", "178.576-15")
test_project_part_tuple_2 = ("1111E.001236", "191.674-01")
test_project_part_tuple_3 = ("1111E.001236", "229.847-00")
test_project_part_tuple_4 = ("1111E.001236", "fake_part_number")

# test project/part tuple
test_project_part_list_tuple_1 = ()
# test_vendors
test_vendor_1 = ""
test_vendor_2 = "48200025"
test_vendor_3 = "fake_vendor"

# test project partlist tuple
test_project_part_list_tuple_1 = ("1111E.001169", ["935.085-00", "935.085-10"])

# test project/vendor/part_list tuple
test_project_vendor_parts_tuple_1 = ("1111E.001169", "48200041",
                                     ["935.085-00", "935.085-10"])

test_project_vendor_parts_tuple_2 = ("1361P.000054", "49100612",
                                     ["187.119-00", "187.120-00", "187.121-00"])

# test project/part/vendor tuple
test_project_part_vendor_tuple = ("1111E.001169", "935.085-00", "48200041")


# ALL_PROJECT_LIST = sql.get_all_project_list("1111E.001236", )

def test_show_cwd():
    cwd = os.getcwd()
    print(">>> cwd: ", cwd)


def test_get_project_info():
    assert sql.get_project_info(test_project_1)['project'] == test_project_1
    assert sql.get_project_info(test_project_blank) is None
    assert sql.get_project_info(test_project_fake) is None
    assert sql.get_project_info(test_project_none) is None
    print(sql.get_project_info(test_project_1).keys())
    print("sop date: ", sql.get_project_info(test_project_1)['sop'])


def test_get_part_list_by_project():
    print(sql.get_part_list_by_project(test_project_1))
    print(sql.get_part_list_by_project(test_project_blank))
    print(sql.get_part_list_by_project(test_project_fake))
    print(sql.get_part_list_by_project(test_project_none))


def test_get_part_general_info():
    print(
        sql.get_part_general_info(
            *test_project_part_tuple_1)["plant_short_name"])
    print(sql.get_part_general_info(*test_project_part_tuple_1)["plant"])
    print(sql.get_part_general_info(*test_project_part_tuple_1)["mgs"])
    print(sql.get_part_general_info(*test_project_part_tuple_1)["mgm"])


def test_get_part_volume_avg():
    print(">>> volume avg ",
          sql.get_part_volume_avg(*test_project_part_tuple_1))
    print(">>> volume avg ",
          sql.get_part_volume_avg(*test_project_part_tuple_2))
    print(">>> volume avg ",
          sql.get_part_volume_avg(*test_project_part_tuple_3))
    print(">>> volume avg ",
          sql.get_part_volume_avg(*test_project_part_tuple_4))


def test_get_part_target_pvo_part():
    print(">>>part_target_pvo_part ",
          sql.get_part_target_pvo_part(*test_project_part_tuple_1))
    print(">>>part_target_pvo_part ",
          sql.get_part_target_pvo_part(*test_project_part_tuple_2))
    print(">>>part_target_pvo_part ",
          sql.get_part_target_pvo_part(*test_project_part_tuple_3))
    print(">>>part_target_pvo_part ",
          sql.get_part_target_pvo_part(*test_project_part_tuple_4))


def test_get_part_target_pvo_investment():
    print(">>>part_target_pvo_investment ",
          sql.get_part_target_pvo_investment(*test_project_part_tuple_1))
    print(">>>part_target_pvo_investment ",
          sql.get_part_target_pvo_investment(*test_project_part_tuple_2))
    print(">>>part_target_pvo_investment ",
          sql.get_part_target_pvo_investment(*test_project_part_tuple_3))
    print(">>>part_target_pvo_investment ",
          sql.get_part_target_pvo_investment(*test_project_part_tuple_4))


def test_get_part_target_pvo_total():
    print(">>>part_target_pvo_total ",
          sql.get_part_target_pvo_total(*test_project_part_tuple_1))
    print(">>>part_target_pvo_total ",
          sql.get_part_target_pvo_total(*test_project_part_tuple_2))
    print(">>>part_target_pvo_total ",
          sql.get_part_target_pvo_total(*test_project_part_tuple_3))
    print(">>>part_target_pvo_total ",
          sql.get_part_target_pvo_total(*test_project_part_tuple_4))


def test_get_part_lifetime():
    print(">>> part lifetime ",
          sql.get_part_lifetime(*test_project_part_tuple_1))
    print(">>> part lifetime ",
          sql.get_part_lifetime(*test_project_part_tuple_2))
    print(">>> part lifetime ",
          sql.get_part_lifetime(*test_project_part_tuple_3))
    print(">>> part lifetime ",
          sql.get_part_lifetime(*test_project_part_tuple_4))


def test_get_target_avg_100EUR():
    print(">>> target avg price ",
          sql.get_part_target_price_avg_100EUR(*test_project_part_tuple_1))
    print(">>> target avg price ",
          sql.get_part_target_price_avg_100EUR(*test_project_part_tuple_2))
    print(">>> target avg price ",
          sql.get_part_target_price_avg_100EUR(*test_project_part_tuple_3))
    print(">>> target avg price ",
          sql.get_part_target_price_avg_100EUR(*test_project_part_tuple_4))


def test_get_part_timing():
    print(">>>part timing ",
          sql.get_part_timing(*test_project_part_tuple_2).keys())
    print(">>>part timing ",
          sql.get_part_timing(*test_project_part_tuple_2)['ppap_date'])
    ppap = sql.get_part_timing(*test_project_part_tuple_2)['ppap_date']
    print("type of ppap: ", type(ppap))


def test_get_vendor_info():
    print(">>>vendor info", sql.get_vendor_info(test_vendor_2).keys())
    print(">>>vendor info", sql.get_vendor_info(test_vendor_2)['vendor_name'])
    print(">>>vendor info",
          sql.get_vendor_info(test_vendor_2)['quality_mgnt_signed'])


def test_get_part_volume_yearly():
    print(">>>volume yearly: ",
          sql.get_part_volume_yearly(*test_project_part_tuple_1))
    print(">>>volume yearly: ",
          sql.get_part_volume_yearly(*test_project_part_tuple_2))
    print(">>>volume yearly: ",
          sql.get_part_volume_yearly(*test_project_part_tuple_3))
    print(">>>volume yearly: ",
          sql.get_part_volume_yearly(*test_project_part_tuple_4))


def test_get_part_price_yearly():
    print(">>>price yearly: ",
          sql.get_part_price_yearly(*test_project_part_vendor_tuple))


def test_project_sop_eop():
    print(">>>project sop/eop ", sql.get_project_sop_eop(test_project_1))
    print(">>>project sop/eop ", sql.get_project_sop_eop(test_project_fake))
    print(">>>project sop/eop ", sql.get_project_sop_eop(test_project_blank))
    print(">>>project sop/eop ", sql.get_project_sop_eop(test_project_none))


# test injection function
def test_xls_inject_risk_eval():
    xls_inject.xls_inject_risk_eval(*test_project_part_list_tuple_1)


def test_xls_inject_cbd_single():
    # file prep
    TEMPLATE_PATH = constant.TEMPLATE_PATH
    file_name = 'cbd'
    file_path = TEMPLATE_PATH + file_name + '.xlsx'

    # load workbook into openpyxl
    wb = load_workbook(file_path)

    # run injection
    xls_inject.xls_inject_cbd_single(*test_project_part_tuple_2, "48200041",
                                     wb)


def test_xls_inject_cbd_project():
    xls_inject.xls_inject_cbd_project(test_project_1)


def test_xls_inject_ss_single():
    # file prep
    TEMPLATE_PATH = constant.TEMPLATE_PATH
    file_name = 'supplier_selection'
    file_path = TEMPLATE_PATH + file_name + '.xlsx'

    # load workbook into openpyxl
    wb = load_workbook(file_path)

    # run injection
    xls_inject.xls_inject_ss_single(*test_project_part_tuple_2, wb)


def test_xls_inject_ss_project():
    xls_inject.xls_inject_ss_project(test_project_1)


# def test_xls_inject_sb():
#     xls_inject.xls_inject_sb(*test_project_part_list_tuple_1)


def test_get_part_volume_weekly():
    print(">>> weekly volume: ",
          sql.get_part_volume_weekly(*test_project_part_vendor_tuple))


def test_get_nl_tool_info():
    rows = sql.get_nl_tool_info(*test_project_vendor_parts_tuple_2)
    for row in rows:
        print(">>> tool info ")
        print(">>> part info", row['part'], ' ', row['tool'], ' ', row['tool_description'])


def test_get_nl_invest_info():
    rows = sql.get_nl_invest_info(*test_project_vendor_parts_tuple_2)
    for row in rows:
        print(">>> invest info ")
        print(">>> part info", row['part'], ' ', row['invest_name'], ' ', row['cost'])


def test_get_project_vendor_qs_yearly():
    print(">>> yearly qs: ", sql.get_project_vendor_qs_yearly(*project_vendor_tuple1))
    print(">>> yearly qs: ", sql.get_project_vendor_qs_yearly(*project_vendor_tuple2))


def test_assemble_nl_info():
    print(">>> nl info: ",
          sql.assemble_nl_info(*test_project_vendor_parts_tuple_1))


def test_generate_nl():
    word.generate_nl(*test_project_vendor_parts_tuple_1)


def test_get_all_project_vendor_tuple():
    print(">>> all project/vendor tuple: ",
          sql.get_all_project_vendor_tuple())


def test_all_project_vendor_parts_tuple_for_nl_info():
    for (project, vendor) in sql.get_all_project_vendor_tuple():
        part_list = sql.get_part_list_by_project_vendor(project, vendor)
        print(">>> nl info: ",
              sql.assemble_nl_info(project, vendor, part_list))


def test_all_project_vendor_parts_tuple_for_nl_generation():
    for (project, vendor) in sql.get_all_project_vendor_tuple():
        part_list = sql.get_part_list_by_project_vendor(project, vendor)
        print(">>> project: ", project)
        print(">>> vendor: ", vendor)
        print(">>> part_list: ", part_list)
        word.generate_nl(project, vendor, part_list)
