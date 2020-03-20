import zipfile
import os
from openpyxl import load_workbook

import config
from app.models import sql_NRM
from config import DOWNLOAD_FOLDER, TEMPLATE_FOLDER

CBD_SHEET_PASSWORD = config.EXCEL_PASSWORD


def xls_inject_risk_eval(project, part_list):
    template_filename = 'risk_eval.xlsx'
    template_path = os.path.join(TEMPLATE_FOLDER, template_filename)
    template_sheet_name = 'Risk Evaluation'

    # load workbook into openpyxl
    wb = load_workbook(template_path)

    # load the sheet
    sheet = wb[template_sheet_name]

    # inject project information
    project_info = sql_NRM.get_project_data_and_info(project)
    if project_info:
        sheet['E1'] = project_info['project_name']

    # inject part information part by part
    n = 0
    for part in part_list:
        print("----debug---xls_inject--- inject part: ", part)
        context = (project, part)
        column = (6 + 2 * n)

        if project_info:
            sheet.cell(row=12, column=column).value = project_info['sop'][3:]

        # part general info
        part_general_info = sql_NRM.get_part_general_info(*context)
        if part_general_info:
            sheet.cell(row=4, column=column).value = part_general_info['part']
            sheet.cell(
                row=5,
                column=column).value = part_general_info['part_description']
            sheet.cell(row=6,
                       column=column).value = part_general_info['mtl_group']
            sheet.cell(
                row=10,
                column=column).value = part_general_info['plant_short_name']
            sheet.cell(row=13,
                       column=column).value = part_general_info['raw_mtl']
            sheet.cell(row=14, column=column).value = part_general_info['mgm']
            sheet.cell(row=15, column=column).value = part_general_info['mgs']
            sheet.cell(row=16,
                       column=column).value = part_general_info['buyer']

        # volume_avg
        vol_avg = sql_NRM.get_part_volume_avg(*context)
        if vol_avg:
            sheet.cell(row=7, column=column).value = int(vol_avg)

        # target price:
        target_price = sql_NRM.get_part_target_price_avg_100EUR(*context)
        sheet.cell(row=8,
                   column=column).value = target_price  # Input as None is OK

        # lifetime:
        lifetime = sql_NRM.get_part_lifetime(*context)
        if lifetime:  # avoid lifetime=0
            sheet.cell(row=9, column=column).value = lifetime

        # part_timing
        part_timing = sql_NRM.get_part_timing(*context)
        if part_timing:
            ppap = part_timing['ppap_date']
            if ppap:
                sheet.cell(row=11,
                           column=column).value = ppap[5:7] + '.' + ppap[0:4]

        # next part
        n += 1

    # save the inject
    output_filename = f"Risk_Evaluation_{project}.xlsx"
    output_file_path = os.path.join(DOWNLOAD_FOLDER, output_filename)
    # remove file is already exists
    if os.path.exists(output_file_path):
        os.remove(output_file_path)
    wb.save(output_file_path)
    print("------debug------ file saved to downloads: ", output_file_path)
    return output_filename


def xls_inject_cbd_single(project, part, vendor, workbook):
    """create single CBD file"""

    # load the sheet
    sheet_name = 'CBD Summary'
    sheet = workbook[sheet_name]

    # part general info
    part_general_info = sql_NRM.get_part_general_info(project, part)
    if part_general_info:
        sheet['D5'] = part_general_info['part']
        sheet['I5'] = part_general_info['part_description']
        sheet['L11'] = part_general_info['mtl_group']

    # vendor info
    sheet['L7'] = vendor
    vendor_info = sql_NRM.get_vendor_info(vendor)
    if vendor_info:
        sheet['O7'] = vendor_info['vendor_name']

    # project info
    project_info = sql_NRM.get_project_data_and_info(project)
    if project_info:
        sheet['I9'] = project_info['project']
        sheet['D9'] = project_info['project_name']
        sheet['I11'] = project_info['sop']

        # use sop year. First check if is digit
        sop_year = project_info['sop'][6:10]
        if sop_year.isdigit():
            sheet['E15'] = int(sop_year)

    # segment (pass)

    # yearly volume
    volume_dict = sql_NRM.get_part_volume_yearly(project, part)
    if volume_dict:
        for i, vol in enumerate(volume_dict):
            if i < 11:
                sheet.cell(row=16, column=5 + i).value = vol

    # save the inject
    output_file_name = f'CBD_{project}_{part}_{vendor}_00.xlsx'
    output_file_path = os.path.join(DOWNLOAD_FOLDER, output_file_name)

    ws = workbook.active
    ws.protection.password = CBD_SHEET_PASSWORD
    ws.protection.enable()
    workbook.save(output_file_path)

    return output_file_path


def xls_inject_cbd_project(project, part_list):
    """downloads CBD in zip file"""
    # file prep
    template_name = 'cbd.xlsx'
    template_path = os.path.join(TEMPLATE_FOLDER, template_name)

    # load workbook into openpyxl
    wb = load_workbook(template_path)

    # helper prep
    output_path_list = []

    for part in part_list:
        # prepare vendor list
        vendor_list = sql_NRM.get_vendor_selected_list_by_project_and_part(project, part)

        for vendor in vendor_list:
            # run the injection and return the file name
            output_file_name = xls_inject_cbd_single(project, part, vendor, wb)
            output_path_list.append(output_file_name)

    # zip the downloads files
    output_zip_name = f"CBD_Package_{project}.zip"
    zip_path = os.path.join(DOWNLOAD_FOLDER, output_zip_name)
    with zipfile.ZipFile(zip_path, 'w') as new_zip:
        for name in output_path_list:
            new_zip.write(name, arcname=name.split('/')[-1])

    # remove the excel files
    for name in output_path_list:
        os.remove(name)

    return output_zip_name


def xls_inject_ss_project(project, part_list):
    """downloads supplier selection in zip file"""

    # file prep (name and template loading)
    template_name = 'ss.xlsx'
    file_path = os.path.join(TEMPLATE_FOLDER, template_name)

    # load workbook into openpyxl
    wb = load_workbook(file_path)

    # helper list
    output_path_list = []

    for part in part_list:
        # run the injection and return the file name
        output_ss_path = xls_inject_ss_single(project, part, wb)
        print(f"[build SS] SS output filename: ", output_ss_path)
        output_path_list.append(output_ss_path)

    print(f"[build SS] SS output file list: ", output_path_list)

    # zip the downloads files
    output_zip_name = f"Supplier_Selection_Package_{project}.zip"
    zip_path = os.path.join(DOWNLOAD_FOLDER, output_zip_name)
    with zipfile.ZipFile(zip_path, 'w') as new_zip:
        for name in output_path_list:
            new_zip.write(name, arcname=name.split('/')[-1])

    # remove the excel files
    for name in output_path_list:
        os.remove(name)

    return output_zip_name


def xls_inject_ss_single(project, part, wb):
    # load the sheet
    sheet_name = 'Supplier Selection'
    sheet = wb[sheet_name]

    # part general info
    part_general_info = sql_NRM.get_part_general_info(project, part)
    if part_general_info:
        sheet['E8'] = part_general_info['part']
        sheet['K8'] = part_general_info['part_description']
        sheet['K10'] = part_general_info['mtl_group']
        sheet['E16'] = part_general_info['mgm']
        sheet['E10'] = part_general_info['cmd_group']
        sheet['K16'] = part_general_info['mgs']
        sheet['E18'] = part_general_info['buyer']
        sheet['K18'] = part_general_info['plant_short_name']
        sheet['E20'] = part_general_info['raw_mtl']
        sheet['R23'] = part_general_info['risk_level']

        # signature name
        sheet['C56'] = part_general_info['buyer']
        sheet['I56'] = part_general_info['mgs']
        sheet['O56'] = part_general_info['sqe']

    # project info
    project_info = sql_NRM.get_project_data_and_info(project)
    if project_info:
        sheet['K6'] = project_info['project']
        sheet['E6'] = project_info['project_name']
        sheet['E12'] = project_info['sop']

    # lifetime:
    lifetime = sql_NRM.get_part_lifetime(project, part)
    if lifetime:  # use if to avoid lifetime=0
        sheet['K12'] = lifetime

    # volume_average
    vol_avg = sql_NRM.get_part_volume_avg(project, part)
    if vol_avg:
        sheet['E14'] = int(vol_avg)

    # PVO
    pvo = sql_NRM.get_part_target_pvo_part(project, part)
    if pvo:
        sheet['K14'] = pvo

    # part_timing
    part_timing = sql_NRM.get_part_timing(project, part)
    if part_timing:
        sheet['R25'] = part_timing['nomination_date']

    # prepare vendor list
    n = 0  # for the row adding
    vendor_list = sql_NRM.get_vendor_selected_list_by_project_and_part(project, part)
    for vendor in vendor_list:

        # inject vendor ID
        sheet.cell(row=29 + n, column=2).value = vendor
        # inject vendor info
        vendor_info = sql_NRM.get_vendor_info(vendor)
        if vendor_info:
            sheet.cell(row=29 + n, column=4).value = vendor_info['vendor_name']
            sheet.cell(row=29 + n, column=7).value = vendor_info['rating']
            sheet.cell(row=29 + n, column=7).value = vendor_info['rating']
            sheet.cell(row=29 + n,
                       column=9).value = vendor_info['risk_assessment']
            sheet.cell(row=29 + n,
                       column=11).value = vendor_info['contract_status']
            sheet.cell(row=29 + n, column=13).value = vendor_info['ppm_fy']
            sheet.cell(row=29 + n,
                       column=14).value = vendor_info['released_production']

    # save the inject
    output_file_name = f'Supplier_Selection_{part}.xlsx'
    output_file_path = os.path.join(DOWNLOAD_FOLDER, output_file_name)
    wb.save(output_file_path)

    return output_file_path


def xls_inject_sb(project, part_list):
    '''generate xlxs file according to given project and part list'''

    file_name = 'source_ge'
    file_path = TEMPLATE_FOLDER + file_name + '.xlsx'
    sheet_name = 'Input'

    # load workbook into openpyxl
    wb = load_workbook(file_path)

    # load the sheet
    sheet = wb[sheet_name]

    # start the injection

    # project information
    project_info = sql_NRM.get_project_data_and_info(project)
    sheet['H3'] = project
    if project_info:
        sheet['H4'] = project_info['project_name']
        sheet['H5'] = project_info['customer']
        sheet['W3'] = project_info['dd_location']
        sheet['W4'] = project_info['plant_short_name']
        sheet['AK3'] = project_info['app']
        sheet['AK4'] = project_info['pjm']
        sheet['AK5'] = project_info['md']
        sheet['AK6'] = project_info['sqa']

        # volumes

        ## volume start year. Use sop year. First check if is digit.
        sop_year = project_info['sop'][6:10]
        if sop_year.isdigit():
            sheet['H10'] = int(sop_year)

        ## project yearly volume.
        # TODO after you build the project extra DB, update these code with iteration style
        sheet['H10'] = project_info['year_1_volume']
        sheet['L10'] = project_info['year_2_volume']
        sheet['P10'] = project_info['year_3_volume']
        sheet['T10'] = project_info['year_4_volume']
        sheet['X10'] = project_info['year_5_volume']
        sheet['AB10'] = project_info['year_6_volume']
        sheet['AF10'] = project_info['year_7_volume']
        sheet['AJ10'] = project_info['year_8_volume']
        sheet['AN10'] = project_info['year_9_volume']
        sheet['AR10'] = project_info['year_10_volume']

        # project timing
        sheet['F16'] = project_info['run_rate_date']
        sheet['F17'] = project_info['pv_hella_date']
        sheet['F18'] = project_info['sop']
        sheet['F19'] = project_info['sop_customer_date']

    # Timing single part

    ## part title (part and description)
    part_info_i = 0
    for part in part_list:
        part_timing_column_num = (6 + 6 * part_info_i)

        # part general info
        part_general_info = sql_NRM.get_part_general_info(project, part)
        if part_general_info and part_info_i < 4:
            sheet.cell(row=24, column=part_timing_column_num
                       ).value = part_general_info['part']
            sheet.cell(row=25, column=part_timing_column_num
                       ).value = part_general_info['part_description']
        # part_timing
        part_timing = sql_NRM.get_part_timing(project, part)
        if part_timing:
            sheet.cell(row=30, column=part_timing_column_num
                       ).value = part_timing['nomination_date']
            # TODO FOT timing missing
            # TODO c-sample timing missing
            sheet.cell(
                row=33,
                column=part_timing_column_num).value = part_timing['ppap_date']

        part_info_i += 1

    # BOM cost section
    bom_part_row_i = 0
    for part in part_list:
        bom_part_row_num = 43 + bom_part_row_i

        ## part general info
        part_general_info = sql_NRM.get_part_general_info(project, part)
        if part_general_info and bom_part_row_i < 4:
            sheet.cell(row=bom_part_row_num,
                       column=2).value = part_general_info['part']
            sheet.cell(row=bom_part_row_num,
                       column=6).value = part_general_info['part_description']
            sheet.cell(row=bom_part_row_num,
                       column=15).value = part_general_info['currency']
            ## target_PVO
            target_pvo_total = sql_NRM.get_part_target_pvo_total(project, part)
            if target_pvo_total:
                sheet.cell(row=bom_part_row_num,
                           column=15).value = target_pvo_total

            ##TODO actual PVO is left at the moment. Which supplier to choose from?

    bom_part_row_i += 1

    # Concept: Single Part Status
    concept_single_part_row_i = 0
    for part in part_list:
        concept_single_part_row_num = 62 + 4 * concept_single_part_row_i
        concept_single_part_col_num = 15

        ## part general info
        part_general_info = sql_NRM.get_part_general_info(project, part)
        if part_general_info and concept_single_part_row_i < 4:
            sheet.cell(row=concept_single_part_row_num,
                       column=15).value = part_general_info['mtl_group']
            sheet.cell(row=concept_single_part_row_num,
                       column=25).value = part_general_info['currency']
        ## target PVO
        target_pvo_total = sql_NRM.get_part_target_pvo_total(project, part)
        if target_pvo_total:
            sheet.cell(row=concept_single_part_row_num,
                       column=20).value = target_pvo_total

        ## sourcing plan date
        part_timing = sql_NRM.get_part_timing(project, part)
        if part_timing:
            sheet.cell(row=concept_single_part_row_num,
                       column=33).value = part_timing['nomination_date']

        concept_single_part_row_i += 1

    # part_1_vendor_1 = project_dict['parts']['part_1']['quotations'][
    #     'vendor_1']['vendor']
    # part_1_vendor_2 = project_dict['parts']['part_1']['quotations'][
    #     'vendor_2']['vendor']
    # part_1_vendor_3 = project_dict['parts']['part_1']['quotations'][
    #     'vendor_3']['vendor']
    # part_1_vendor_4 = project_dict['parts']['part_1']['quotations'][
    #     'vendor_4']['vendor']
    # part_1_vendor_5 = project_dict['parts']['part_1']['quotations'][
    #     'vendor_5']['vendor']
    # part_1_vendor_6 = project_dict['parts']['part_1']['quotations'][
    #     'vendor_6']['vendor']
    # part_2_vendor_1 = project_dict['parts']['part_2']['quotations'][
    #     'vendor_1']['vendor']
    # part_2_vendor_2 = project_dict['parts']['part_2']['quotations'][
    #     'vendor_2']['vendor']
    # part_2_vendor_3 = project_dict['parts']['part_2']['quotations'][
    #     'vendor_3']['vendor']
    # part_2_vendor_4 = project_dict['parts']['part_2']['quotations'][
    #     'vendor_4']['vendor']
    # part_2_vendor_5 = project_dict['parts']['part_2']['quotations'][
    #     'vendor_5']['vendor']
    # part_2_vendor_6 = project_dict['parts']['part_2']['quotations'][
    #     'vendor_6']['vendor']
    # part_3_vendor_1 = project_dict['parts']['part_3']['quotations'][
    #     'vendor_1']['vendor']
    # part_3_vendor_2 = project_dict['parts']['part_3']['quotations'][
    #     'vendor_2']['vendor']
    # part_3_vendor_3 = project_dict['parts']['part_3']['quotations'][
    #     'vendor_3']['vendor']
    # part_3_vendor_4 = project_dict['parts']['part_3']['quotations'][
    #     'vendor_4']['vendor']
    # part_3_vendor_5 = project_dict['parts']['part_3']['quotations'][
    #     'vendor_5']['vendor']
    # part_3_vendor_6 = project_dict['parts']['part_3']['quotations'][
    #     'vendor_6']['vendor']
    # part_4_vendor_1 = project_dict['parts']['part_4']['quotations'][
    #     'vendor_1']['vendor']
    # part_4_vendor_2 = project_dict['parts']['part_4']['quotations'][
    #     'vendor_2']['vendor']
    # part_4_vendor_3 = project_dict['parts']['part_4']['quotations'][
    #     'vendor_3']['vendor']
    # part_4_vendor_4 = project_dict['parts']['part_4']['quotations'][
    #     'vendor_4']['vendor']
    # part_4_vendor_5 = project_dict['parts']['part_4']['quotations'][
    #     'vendor_5']['vendor']
    # part_4_vendor_6 = project_dict['parts']['part_4']['quotations'][
    #     'vendor_6']['vendor']

    # # supplier names
    # sheet['I80'] = project_dict['vendors'][part_1_vendor_1][
    #     'short_name'] if part_1_vendor_1 else None
    # sheet['P80'] = project_dict['vendors'][part_1_vendor_2][
    #     'short_name'] if part_1_vendor_2 else None
    # sheet['W80'] = project_dict['vendors'][part_1_vendor_3][
    #     'short_name'] if part_1_vendor_3 else None
    # sheet['AD80'] = project_dict['vendors'][part_1_vendor_4][
    #     'short_name'] if part_1_vendor_4 else None
    # sheet['AK80'] = project_dict['vendors'][part_1_vendor_5][
    #     'short_name'] if part_1_vendor_5 else None
    # sheet['AR80'] = project_dict['vendors'][part_1_vendor_6][
    #     'short_name'] if part_1_vendor_6 else None
    # sheet['BH80'] = project_dict['vendors'][part_2_vendor_1][
    #     'short_name'] if part_2_vendor_1 else None
    # sheet['BO80'] = project_dict['vendors'][part_2_vendor_2][
    #     'short_name'] if part_2_vendor_2 else None
    # sheet['BV80'] = project_dict['vendors'][part_2_vendor_3][
    #     'short_name'] if part_2_vendor_3 else None
    # sheet['CC80'] = project_dict['vendors'][part_2_vendor_4][
    #     'short_name'] if part_2_vendor_4 else None
    # sheet['CJ80'] = project_dict['vendors'][part_2_vendor_5][
    #     'short_name'] if part_2_vendor_5 else None
    # sheet['CQ80'] = project_dict['vendors'][part_2_vendor_6][
    #     'short_name'] if part_2_vendor_6 else None
    # sheet['I89'] = project_dict['vendors'][part_3_vendor_1][
    #     'short_name'] if part_3_vendor_1 else None
    # sheet['P89'] = project_dict['vendors'][part_3_vendor_2][
    #     'short_name'] if part_3_vendor_2 else None
    # sheet['W89'] = project_dict['vendors'][part_3_vendor_3][
    #     'short_name'] if part_3_vendor_3 else None
    # sheet['AD89'] = project_dict['vendors'][part_3_vendor_4][
    #     'short_name'] if part_3_vendor_4 else None
    # sheet['AK89'] = project_dict['vendors'][part_3_vendor_5][
    #     'short_name'] if part_3_vendor_5 else None
    # sheet['AR89'] = project_dict['vendors'][part_3_vendor_6][
    #     'short_name'] if part_3_vendor_6 else None
    # sheet['BH89'] = project_dict['vendors'][part_4_vendor_1][
    #     'short_name'] if part_4_vendor_1 else None
    # sheet['BO89'] = project_dict['vendors'][part_4_vendor_2][
    #     'short_name'] if part_4_vendor_2 else None
    # sheet['BV89'] = project_dict['vendors'][part_4_vendor_3][
    #     'short_name'] if part_4_vendor_3 else None
    # sheet['CC89'] = project_dict['vendors'][part_4_vendor_4][
    #     'short_name'] if part_4_vendor_4 else None
    # sheet['CJ89'] = project_dict['vendors'][part_4_vendor_5][
    #     'short_name'] if part_4_vendor_5 else None
    # sheet['CQ89'] = project_dict['vendors'][part_4_vendor_6][
    #     'short_name'] if part_4_vendor_6 else None

    # # supplier rating
    # sheet['I81'] = project_dict['vendors'][part_1_vendor_1][
    #     'rating'] if part_1_vendor_1 else None
    # sheet['P81'] = project_dict['vendors'][part_1_vendor_2][
    #     'rating'] if part_1_vendor_2 else None
    # sheet['W81'] = project_dict['vendors'][part_1_vendor_3][
    #     'rating'] if part_1_vendor_3 else None
    # sheet['AD81'] = project_dict['vendors'][part_1_vendor_4][
    #     'rating'] if part_1_vendor_4 else None
    # sheet['AK81'] = project_dict['vendors'][part_1_vendor_5][
    #     'rating'] if part_1_vendor_5 else None
    # sheet['AR81'] = project_dict['vendors'][part_1_vendor_6][
    #     'rating'] if part_1_vendor_6 else None
    # sheet['BH81'] = project_dict['vendors'][part_2_vendor_1][
    #     'rating'] if part_2_vendor_1 else None
    # sheet['BO81'] = project_dict['vendors'][part_2_vendor_2][
    #     'rating'] if part_2_vendor_2 else None
    # sheet['BV81'] = project_dict['vendors'][part_2_vendor_3][
    #     'rating'] if part_2_vendor_3 else None
    # sheet['CC81'] = project_dict['vendors'][part_2_vendor_4][
    #     'rating'] if part_2_vendor_4 else None
    # sheet['CJ81'] = project_dict['vendors'][part_2_vendor_5][
    #     'rating'] if part_2_vendor_5 else None
    # sheet['CQ81'] = project_dict['vendors'][part_2_vendor_6][
    #     'rating'] if part_2_vendor_6 else None
    # sheet['I90'] = project_dict['vendors'][part_3_vendor_1][
    #     'rating'] if part_3_vendor_1 else None
    # sheet['P90'] = project_dict['vendors'][part_3_vendor_2][
    #     'rating'] if part_3_vendor_2 else None
    # sheet['W90'] = project_dict['vendors'][part_3_vendor_3][
    #     'rating'] if part_3_vendor_3 else None
    # sheet['AD90'] = project_dict['vendors'][part_3_vendor_4][
    #     'rating'] if part_3_vendor_4 else None
    # sheet['AK90'] = project_dict['vendors'][part_3_vendor_5][
    #     'rating'] if part_3_vendor_5 else None
    # sheet['AR90'] = project_dict['vendors'][part_3_vendor_6][
    #     'rating'] if part_3_vendor_6 else None
    # sheet['BH90'] = project_dict['vendors'][part_4_vendor_1][
    #     'rating'] if part_4_vendor_1 else None
    # sheet['BO90'] = project_dict['vendors'][part_4_vendor_2][
    #     'rating'] if part_4_vendor_2 else None
    # sheet['BV90'] = project_dict['vendors'][part_4_vendor_3][
    #     'rating'] if part_4_vendor_3 else None
    # sheet['CC90'] = project_dict['vendors'][part_4_vendor_4][
    #     'rating'] if part_4_vendor_4 else None
    # sheet['CJ90'] = project_dict['vendors'][part_4_vendor_5][
    #     'rating'] if part_4_vendor_5 else None
    # sheet['CQ90'] = project_dict['vendors'][part_4_vendor_6][
    #     'rating'] if part_4_vendor_6 else None

    # # supplier contract status
    # sheet['I82'] = project_dict['vendors'][part_1_vendor_1][
    #     'contract_status'] if part_1_vendor_1 else None
    # sheet['P82'] = project_dict['vendors'][part_1_vendor_2][
    #     'contract_status'] if part_1_vendor_2 else None
    # sheet['W82'] = project_dict['vendors'][part_1_vendor_3][
    #     'contract_status'] if part_1_vendor_3 else None
    # sheet['AD82'] = project_dict['vendors'][part_1_vendor_4][
    #     'contract_status'] if part_1_vendor_4 else None
    # sheet['AK82'] = project_dict['vendors'][part_1_vendor_5][
    #     'contract_status'] if part_1_vendor_5 else None
    # sheet['AR82'] = project_dict['vendors'][part_1_vendor_6][
    #     'contract_status'] if part_1_vendor_6 else None
    # sheet['BH82'] = project_dict['vendors'][part_2_vendor_1][
    #     'contract_status'] if part_2_vendor_1 else None
    # sheet['BO82'] = project_dict['vendors'][part_2_vendor_2][
    #     'contract_status'] if part_2_vendor_2 else None
    # sheet['BV82'] = project_dict['vendors'][part_2_vendor_3][
    #     'contract_status'] if part_2_vendor_3 else None
    # sheet['CC82'] = project_dict['vendors'][part_2_vendor_4][
    #     'contract_status'] if part_2_vendor_4 else None
    # sheet['CJ82'] = project_dict['vendors'][part_2_vendor_5][
    #     'contract_status'] if part_2_vendor_5 else None
    # sheet['CQ82'] = project_dict['vendors'][part_2_vendor_6][
    #     'contract_status'] if part_2_vendor_6 else None
    # sheet['I91'] = project_dict['vendors'][part_3_vendor_1][
    #     'contract_status'] if part_3_vendor_1 else None
    # sheet['P91'] = project_dict['vendors'][part_3_vendor_2][
    #     'contract_status'] if part_3_vendor_2 else None
    # sheet['W91'] = project_dict['vendors'][part_3_vendor_3][
    #     'contract_status'] if part_3_vendor_3 else None
    # sheet['AD91'] = project_dict['vendors'][part_3_vendor_4][
    #     'contract_status'] if part_3_vendor_4 else None
    # sheet['AK91'] = project_dict['vendors'][part_3_vendor_5][
    #     'contract_status'] if part_3_vendor_5 else None
    # sheet['AR91'] = project_dict['vendors'][part_3_vendor_6][
    #     'contract_status'] if part_3_vendor_6 else None
    # sheet['BH91'] = project_dict['vendors'][part_4_vendor_1][
    #     'contract_status'] if part_4_vendor_1 else None
    # sheet['BO91'] = project_dict['vendors'][part_4_vendor_2][
    #     'contract_status'] if part_4_vendor_2 else None
    # sheet['BV91'] = project_dict['vendors'][part_4_vendor_3][
    #     'contract_status'] if part_4_vendor_3 else None
    # sheet['CC91'] = project_dict['vendors'][part_4_vendor_4][
    #     'contract_status'] if part_4_vendor_4 else None
    # sheet['CJ91'] = project_dict['vendors'][part_4_vendor_5][
    #     'contract_status'] if part_4_vendor_5 else None
    # sheet['CQ91'] = project_dict['vendors'][part_4_vendor_6][
    #     'contract_status'] if part_4_vendor_6 else None

    # # supplier contract status
    # sheet['I83'] = project_dict['vendors'][part_1_vendor_1][
    #     'ppm_global'] if part_1_vendor_1 else None
    # sheet['P83'] = project_dict['vendors'][part_1_vendor_2][
    #     'ppm_global'] if part_1_vendor_2 else None
    # sheet['W83'] = project_dict['vendors'][part_1_vendor_3][
    #     'ppm_global'] if part_1_vendor_3 else None
    # sheet['AD83'] = project_dict['vendors'][part_1_vendor_4][
    #     'ppm_global'] if part_1_vendor_4 else None
    # sheet['AK83'] = project_dict['vendors'][part_1_vendor_5][
    #     'ppm_global'] if part_1_vendor_5 else None
    # sheet['AR83'] = project_dict['vendors'][part_1_vendor_6][
    #     'ppm_global'] if part_1_vendor_6 else None
    # sheet['BH83'] = project_dict['vendors'][part_2_vendor_1][
    #     'ppm_global'] if part_2_vendor_1 else None
    # sheet['BO83'] = project_dict['vendors'][part_2_vendor_2][
    #     'ppm_global'] if part_2_vendor_2 else None
    # sheet['BV83'] = project_dict['vendors'][part_2_vendor_3][
    #     'ppm_global'] if part_2_vendor_3 else None
    # sheet['CC83'] = project_dict['vendors'][part_2_vendor_4][
    #     'ppm_global'] if part_2_vendor_4 else None
    # sheet['CJ83'] = project_dict['vendors'][part_2_vendor_5][
    #     'ppm_global'] if part_2_vendor_5 else None
    # sheet['CQ83'] = project_dict['vendors'][part_2_vendor_6][
    #     'ppm_global'] if part_2_vendor_6 else None
    # sheet['I92'] = project_dict['vendors'][part_3_vendor_1][
    #     'ppm_global'] if part_3_vendor_1 else None
    # sheet['P92'] = project_dict['vendors'][part_3_vendor_2][
    #     'ppm_global'] if part_3_vendor_2 else None
    # sheet['W92'] = project_dict['vendors'][part_3_vendor_3][
    #     'ppm_global'] if part_3_vendor_3 else None
    # sheet['AD92'] = project_dict['vendors'][part_3_vendor_4][
    #     'ppm_global'] if part_3_vendor_4 else None
    # sheet['AK92'] = project_dict['vendors'][part_3_vendor_5][
    #     'ppm_global'] if part_3_vendor_5 else None
    # sheet['AR92'] = project_dict['vendors'][part_3_vendor_6][
    #     'ppm_global'] if part_3_vendor_6 else None
    # sheet['BH92'] = project_dict['vendors'][part_4_vendor_1][
    #     'ppm_global'] if part_4_vendor_1 else None
    # sheet['BO92'] = project_dict['vendors'][part_4_vendor_2][
    #     'ppm_global'] if part_4_vendor_2 else None
    # sheet['BV92'] = project_dict['vendors'][part_4_vendor_3][
    #     'ppm_global'] if part_4_vendor_3 else None
    # sheet['CC92'] = project_dict['vendors'][part_4_vendor_4][
    #     'ppm_global'] if part_4_vendor_4 else None
    # sheet['CJ92'] = project_dict['vendors'][part_4_vendor_5][
    #     'ppm_global'] if part_4_vendor_5 else None
    # sheet['CQ92'] = project_dict['vendors'][part_4_vendor_6][
    #     'ppm_global'] if part_4_vendor_6 else None

    # # supplier escalation level
    # sheet['I84'] = project_dict['vendors'][part_1_vendor_1][
    #     'escalation_level'] if part_1_vendor_1 else None
    # sheet['P84'] = project_dict['vendors'][part_1_vendor_2][
    #     'escalation_level'] if part_1_vendor_2 else None
    # sheet['W84'] = project_dict['vendors'][part_1_vendor_3][
    #     'escalation_level'] if part_1_vendor_3 else None
    # sheet['AD84'] = project_dict['vendors'][part_1_vendor_4][
    #     'escalation_level'] if part_1_vendor_4 else None
    # sheet['AK84'] = project_dict['vendors'][part_1_vendor_5][
    #     'escalation_level'] if part_1_vendor_5 else None
    # sheet['AR84'] = project_dict['vendors'][part_1_vendor_6][
    #     'escalation_level'] if part_1_vendor_6 else None
    # sheet['BH84'] = project_dict['vendors'][part_2_vendor_1][
    #     'escalation_level'] if part_2_vendor_1 else None
    # sheet['BO84'] = project_dict['vendors'][part_2_vendor_2][
    #     'escalation_level'] if part_2_vendor_2 else None
    # sheet['BV84'] = project_dict['vendors'][part_2_vendor_3][
    #     'escalation_level'] if part_2_vendor_3 else None
    # sheet['CC84'] = project_dict['vendors'][part_2_vendor_4][
    #     'escalation_level'] if part_2_vendor_4 else None
    # sheet['CJ84'] = project_dict['vendors'][part_2_vendor_5][
    #     'escalation_level'] if part_2_vendor_5 else None
    # sheet['CQ84'] = project_dict['vendors'][part_2_vendor_6][
    #     'escalation_level'] if part_2_vendor_6 else None
    # sheet['I93'] = project_dict['vendors'][part_3_vendor_1][
    #     'escalation_level'] if part_3_vendor_1 else None
    # sheet['P93'] = project_dict['vendors'][part_3_vendor_2][
    #     'escalation_level'] if part_3_vendor_2 else None
    # sheet['W93'] = project_dict['vendors'][part_3_vendor_3][
    #     'escalation_level'] if part_3_vendor_3 else None
    # sheet['AD93'] = project_dict['vendors'][part_3_vendor_4][
    #     'escalation_level'] if part_3_vendor_4 else None
    # sheet['AK93'] = project_dict['vendors'][part_3_vendor_5][
    #     'escalation_level'] if part_3_vendor_5 else None
    # sheet['AR93'] = project_dict['vendors'][part_3_vendor_6][
    #     'escalation_level'] if part_3_vendor_6 else None
    # sheet['BH93'] = project_dict['vendors'][part_4_vendor_1][
    #     'escalation_level'] if part_4_vendor_1 else None
    # sheet['BO93'] = project_dict['vendors'][part_4_vendor_2][
    #     'escalation_level'] if part_4_vendor_2 else None
    # sheet['BV93'] = project_dict['vendors'][part_4_vendor_3][
    #     'escalation_level'] if part_4_vendor_3 else None
    # sheet['CC93'] = project_dict['vendors'][part_4_vendor_4][
    #     'escalation_level'] if part_4_vendor_4 else None
    # sheet['CJ93'] = project_dict['vendors'][part_4_vendor_5][
    #     'escalation_level'] if part_4_vendor_5 else None
    # sheet['CQ93'] = project_dict['vendors'][part_4_vendor_6][
    #     'escalation_level'] if part_4_vendor_6 else None

    # # supplier escalation level
    # sheet['I85'] = project_dict['vendors'][part_1_vendor_1][
    #     'released_production'] if part_1_vendor_1 else None
    # sheet['P85'] = project_dict['vendors'][part_1_vendor_2][
    #     'released_production'] if part_1_vendor_2 else None
    # sheet['W85'] = project_dict['vendors'][part_1_vendor_3][
    #     'released_production'] if part_1_vendor_3 else None
    # sheet['AD85'] = project_dict['vendors'][part_1_vendor_4][
    #     'released_production'] if part_1_vendor_4 else None
    # sheet['AK85'] = project_dict['vendors'][part_1_vendor_5][
    #     'released_production'] if part_1_vendor_5 else None
    # sheet['AR85'] = project_dict['vendors'][part_1_vendor_6][
    #     'released_production'] if part_1_vendor_6 else None
    # sheet['BH85'] = project_dict['vendors'][part_2_vendor_1][
    #     'released_production'] if part_2_vendor_1 else None
    # sheet['BO85'] = project_dict['vendors'][part_2_vendor_2][
    #     'released_production'] if part_2_vendor_2 else None
    # sheet['BV85'] = project_dict['vendors'][part_2_vendor_3][
    #     'released_production'] if part_2_vendor_3 else None
    # sheet['CC85'] = project_dict['vendors'][part_2_vendor_4][
    #     'released_production'] if part_2_vendor_4 else None
    # sheet['CJ85'] = project_dict['vendors'][part_2_vendor_5][
    #     'released_production'] if part_2_vendor_5 else None
    # sheet['CQ85'] = project_dict['vendors'][part_2_vendor_6][
    #     'released_production'] if part_2_vendor_6 else None
    # sheet['I94'] = project_dict['vendors'][part_3_vendor_1][
    #     'released_production'] if part_3_vendor_1 else None
    # sheet['P94'] = project_dict['vendors'][part_3_vendor_2][
    #     'released_production'] if part_3_vendor_2 else None
    # sheet['W94'] = project_dict['vendors'][part_3_vendor_3][
    #     'released_production'] if part_3_vendor_3 else None
    # sheet['AD94'] = project_dict['vendors'][part_3_vendor_4][
    #     'released_production'] if part_3_vendor_4 else None
    # sheet['AK94'] = project_dict['vendors'][part_3_vendor_5][
    #     'released_production'] if part_3_vendor_5 else None
    # sheet['AR94'] = project_dict['vendors'][part_3_vendor_6][
    #     'released_production'] if part_3_vendor_6 else None
    # sheet['BH94'] = project_dict['vendors'][part_4_vendor_1][
    #     'released_production'] if part_4_vendor_1 else None
    # sheet['BO94'] = project_dict['vendors'][part_4_vendor_2][
    #     'released_production'] if part_4_vendor_2 else None
    # sheet['BV94'] = project_dict['vendors'][part_4_vendor_3][
    #     'released_production'] if part_4_vendor_3 else None
    # sheet['CC94'] = project_dict['vendors'][part_4_vendor_4][
    #     'released_production'] if part_4_vendor_4 else None
    # sheet['CJ94'] = project_dict['vendors'][part_4_vendor_5][
    #     'released_production'] if part_4_vendor_5 else None
    # sheet['CQ94'] = project_dict['vendors'][part_4_vendor_6][
    #     'released_production'] if part_4_vendor_6 else None

    # # supplier name again
    # sheet['O100'] = project_dict['vendors'][part_1_vendor_1][
    #     'short_name'] if part_1_vendor_1 else None
    # sheet['V100'] = project_dict['vendors'][part_1_vendor_2][
    #     'short_name'] if part_1_vendor_2 else None
    # sheet['AC100'] = project_dict['vendors'][part_1_vendor_3][
    #     'short_name'] if part_1_vendor_3 else None
    # sheet['AJ100'] = project_dict['vendors'][part_1_vendor_4][
    #     'short_name'] if part_1_vendor_4 else None
    # sheet['BF100'] = project_dict['vendors'][part_2_vendor_1][
    #     'short_name'] if part_2_vendor_1 else None
    # sheet['BM100'] = project_dict['vendors'][part_2_vendor_2][
    #     'short_name'] if part_2_vendor_2 else None
    # sheet['BT100'] = project_dict['vendors'][part_2_vendor_3][
    #     'short_name'] if part_2_vendor_3 else None
    # sheet['CA100'] = project_dict['vendors'][part_2_vendor_4][
    #     'short_name'] if part_2_vendor_4 else None
    # sheet['CW100'] = project_dict['vendors'][part_3_vendor_1][
    #     'short_name'] if part_3_vendor_1 else None
    # sheet['DD100'] = project_dict['vendors'][part_3_vendor_2][
    #     'short_name'] if part_3_vendor_2 else None
    # sheet['DK100'] = project_dict['vendors'][part_3_vendor_3][
    #     'short_name'] if part_3_vendor_3 else None
    # sheet['DR100'] = project_dict['vendors'][part_3_vendor_4][
    #     'short_name'] if part_3_vendor_4 else None
    # sheet['EN100'] = project_dict['vendors'][part_4_vendor_1][
    #     'short_name'] if part_4_vendor_1 else None
    # sheet['EU100'] = project_dict['vendors'][part_4_vendor_2][
    #     'short_name'] if part_4_vendor_2 else None
    # sheet['FB100'] = project_dict['vendors'][part_4_vendor_3][
    #     'short_name'] if part_4_vendor_3 else None
    # sheet['FI100'] = project_dict['vendors'][part_4_vendor_4][
    #     'short_name'] if part_4_vendor_4 else None

    # # supplier framwork agreement
    # sheet['O101'] = 'yes' if part_1_vendor_1 and project_dict['vendors'][
    #     part_1_vendor_1]['framework_date'] else None
    # sheet['V101'] = 'yes' if part_1_vendor_2 and project_dict['vendors'][
    #     part_1_vendor_2]['framework_date'] else None
    # sheet['AC101'] = 'yes' if part_1_vendor_3 and project_dict['vendors'][
    #     part_1_vendor_3]['framework_date'] else None
    # sheet['AJ101'] = 'yes' if part_1_vendor_4 and project_dict['vendors'][
    #     part_1_vendor_4]['framework_date'] else None
    # sheet['BF101'] = 'yes' if part_2_vendor_1 and project_dict['vendors'][
    #     part_2_vendor_1]['framework_date'] else None
    # sheet['BM101'] = 'yes' if part_2_vendor_2 and project_dict['vendors'][
    #     part_2_vendor_2]['framework_date'] else None
    # sheet['BT101'] = 'yes' if part_2_vendor_3 and project_dict['vendors'][
    #     part_2_vendor_3]['framework_date'] else None
    # sheet['CA101'] = 'yes' if part_2_vendor_4 and project_dict['vendors'][
    #     part_2_vendor_4]['framework_date'] else None
    # sheet['CW101'] = 'yes' if part_3_vendor_1 and project_dict['vendors'][
    #     part_3_vendor_1]['framework_date'] else None
    # sheet['DD101'] = 'yes' if part_3_vendor_2 and project_dict['vendors'][
    #     part_3_vendor_2]['framework_date'] else None
    # sheet['DK101'] = 'yes' if part_3_vendor_3 and project_dict['vendors'][
    #     part_3_vendor_3]['framework_date'] else None
    # sheet['DR101'] = 'yes' if part_3_vendor_4 and project_dict['vendors'][
    #     part_3_vendor_4]['framework_date'] else None
    # sheet['EN101'] = 'yes' if part_4_vendor_1 and project_dict['vendors'][
    #     part_4_vendor_1]['framework_date'] else None
    # sheet['EU101'] = 'yes' if part_4_vendor_2 and project_dict['vendors'][
    #     part_4_vendor_2]['framework_date'] else None
    # sheet['FB101'] = 'yes' if part_4_vendor_3 and project_dict['vendors'][
    #     part_4_vendor_3]['framework_date'] else None
    # sheet['FI101'] = 'yes' if part_4_vendor_4 and project_dict['vendors'][
    #     part_4_vendor_4]['framework_date'] else None

    # # supplier delivery regulation
    # sheet['O102'] = 'yes' if part_1_vendor_1 and project_dict['vendors'][
    #     part_1_vendor_1]['delivery_regulation_date'] else None
    # sheet['V102'] = 'yes' if part_1_vendor_2 and project_dict['vendors'][
    #     part_1_vendor_2]['delivery_regulation_date'] else None
    # sheet['AC102'] = 'yes' if part_1_vendor_3 and project_dict['vendors'][
    #     part_1_vendor_3]['delivery_regulation_date'] else None
    # sheet['AJ102'] = 'yes' if part_1_vendor_4 and project_dict['vendors'][
    #     part_1_vendor_4]['delivery_regulation_date'] else None
    # sheet['BF102'] = 'yes' if part_2_vendor_1 and project_dict['vendors'][
    #     part_2_vendor_1]['delivery_regulation_date'] else None
    # sheet['BM102'] = 'yes' if part_2_vendor_2 and project_dict['vendors'][
    #     part_2_vendor_2]['delivery_regulation_date'] else None
    # sheet['BT102'] = 'yes' if part_2_vendor_3 and project_dict['vendors'][
    #     part_2_vendor_3]['delivery_regulation_date'] else None
    # sheet['CA102'] = 'yes' if part_2_vendor_4 and project_dict['vendors'][
    #     part_2_vendor_4]['delivery_regulation_date'] else None
    # sheet['CW102'] = 'yes' if part_3_vendor_1 and project_dict['vendors'][
    #     part_3_vendor_1]['delivery_regulation_date'] else None
    # sheet['DD102'] = 'yes' if part_3_vendor_2 and project_dict['vendors'][
    #     part_3_vendor_2]['delivery_regulation_date'] else None
    # sheet['DK102'] = 'yes' if part_3_vendor_3 and project_dict['vendors'][
    #     part_3_vendor_3]['delivery_regulation_date'] else None
    # sheet['DR102'] = 'yes' if part_3_vendor_4 and project_dict['vendors'][
    #     part_3_vendor_4]['delivery_regulation_date'] else None
    # sheet['EN102'] = 'yes' if part_4_vendor_1 and project_dict['vendors'][
    #     part_4_vendor_1]['delivery_regulation_date'] else None
    # sheet['EU102'] = 'yes' if part_4_vendor_2 and project_dict['vendors'][
    #     part_4_vendor_2]['delivery_regulation_date'] else None
    # sheet['FB102'] = 'yes' if part_4_vendor_3 and project_dict['vendors'][
    #     part_4_vendor_3]['delivery_regulation_date'] else None
    # sheet['FI102'] = 'yes' if part_4_vendor_4 and project_dict['vendors'][
    #     part_4_vendor_4]['delivery_regulation_date'] else None

    # # supplier tool contract
    # sheet['O103'] = 'yes' if part_1_vendor_1 and project_dict['vendors'][
    #     part_1_vendor_1]['tool_contract_date'] else None
    # sheet['V103'] = 'yes' if part_1_vendor_2 and project_dict['vendors'][
    #     part_1_vendor_2]['tool_contract_date'] else None
    # sheet['AC103'] = 'yes' if part_1_vendor_3 and project_dict['vendors'][
    #     part_1_vendor_3]['tool_contract_date'] else None
    # sheet['AJ103'] = 'yes' if part_1_vendor_4 and project_dict['vendors'][
    #     part_1_vendor_4]['tool_contract_date'] else None
    # sheet['BF103'] = 'yes' if part_2_vendor_1 and project_dict['vendors'][
    #     part_2_vendor_1]['tool_contract_date'] else None
    # sheet['BM103'] = 'yes' if part_2_vendor_2 and project_dict['vendors'][
    #     part_2_vendor_2]['tool_contract_date'] else None
    # sheet['BT103'] = 'yes' if part_2_vendor_3 and project_dict['vendors'][
    #     part_2_vendor_3]['tool_contract_date'] else None
    # sheet['CA103'] = 'yes' if part_2_vendor_4 and project_dict['vendors'][
    #     part_2_vendor_4]['tool_contract_date'] else None
    # sheet['CW103'] = 'yes' if part_3_vendor_1 and project_dict['vendors'][
    #     part_3_vendor_1]['tool_contract_date'] else None
    # sheet['DD103'] = 'yes' if part_3_vendor_2 and project_dict['vendors'][
    #     part_3_vendor_2]['tool_contract_date'] else None
    # sheet['DK103'] = 'yes' if part_3_vendor_3 and project_dict['vendors'][
    #     part_3_vendor_3]['tool_contract_date'] else None
    # sheet['DR103'] = 'yes' if part_3_vendor_4 and project_dict['vendors'][
    #     part_3_vendor_4]['tool_contract_date'] else None
    # sheet['EN103'] = 'yes' if part_4_vendor_1 and project_dict['vendors'][
    #     part_4_vendor_1]['tool_contract_date'] else None
    # sheet['EU103'] = 'yes' if part_4_vendor_2 and project_dict['vendors'][
    #     part_4_vendor_2]['tool_contract_date'] else None
    # sheet['FB103'] = 'yes' if part_4_vendor_3 and project_dict['vendors'][
    #     part_4_vendor_3]['tool_contract_date'] else None
    # sheet['FI103'] = 'yes' if part_4_vendor_4 and project_dict['vendors'][
    #     part_4_vendor_4]['tool_contract_date'] else None

    # # supplier quality management
    # sheet['O104'] = project_dict['vendors'][part_1_vendor_1][
    #     'quality_mgnt_signed'] if part_1_vendor_1 else None
    # sheet['V104'] = project_dict['vendors'][part_1_vendor_2][
    #     'quality_mgnt_signed'] if part_1_vendor_2 else None
    # sheet['AC104'] = project_dict['vendors'][part_1_vendor_3][
    #     'quality_mgnt_signed'] if part_1_vendor_3 else None
    # sheet['AJ104'] = project_dict['vendors'][part_1_vendor_4][
    #     'quality_mgnt_signed'] if part_1_vendor_4 else None
    # sheet['BF104'] = project_dict['vendors'][part_2_vendor_1][
    #     'quality_mgnt_signed'] if part_2_vendor_1 else None
    # sheet['BM104'] = project_dict['vendors'][part_2_vendor_2][
    #     'quality_mgnt_signed'] if part_2_vendor_2 else None
    # sheet['BT104'] = project_dict['vendors'][part_2_vendor_3][
    #     'quality_mgnt_signed'] if part_2_vendor_3 else None
    # sheet['CA104'] = project_dict['vendors'][part_2_vendor_4][
    #     'quality_mgnt_signed'] if part_2_vendor_4 else None
    # sheet['CW104'] = project_dict['vendors'][part_3_vendor_1][
    #     'quality_mgnt_signed'] if part_3_vendor_1 else None
    # sheet['DD104'] = project_dict['vendors'][part_3_vendor_2][
    #     'quality_mgnt_signed'] if part_3_vendor_2 else None
    # sheet['DK104'] = project_dict['vendors'][part_3_vendor_3][
    #     'quality_mgnt_signed'] if part_3_vendor_3 else None
    # sheet['DR104'] = project_dict['vendors'][part_3_vendor_4][
    #     'quality_mgnt_signed'] if part_3_vendor_4 else None
    # sheet['EN104'] = project_dict['vendors'][part_4_vendor_1][
    #     'quality_mgnt_signed'] if part_4_vendor_1 else None
    # sheet['EU104'] = project_dict['vendors'][part_4_vendor_2][
    #     'quality_mgnt_signed'] if part_4_vendor_2 else None
    # sheet['FB104'] = project_dict['vendors'][part_4_vendor_3][
    #     'quality_mgnt_signed'] if part_4_vendor_3 else None
    # sheet['FI104'] = project_dict['vendors'][part_4_vendor_4][
    #     'quality_mgnt_signed'] if part_4_vendor_4 else None

    # # supplier logistic guideline
    # sheet['O105'] = project_dict['vendors'][part_1_vendor_1][
    #     'logistic_guideline_signed'] if part_1_vendor_1 else None
    # sheet['V105'] = project_dict['vendors'][part_1_vendor_2][
    #     'logistic_guideline_signed'] if part_1_vendor_2 else None
    # sheet['AC105'] = project_dict['vendors'][part_1_vendor_3][
    #     'logistic_guideline_signed'] if part_1_vendor_3 else None
    # sheet['AJ105'] = project_dict['vendors'][part_1_vendor_4][
    #     'logistic_guideline_signed'] if part_1_vendor_4 else None
    # sheet['BF105'] = project_dict['vendors'][part_2_vendor_1][
    #     'logistic_guideline_signed'] if part_2_vendor_1 else None
    # sheet['BM105'] = project_dict['vendors'][part_2_vendor_2][
    #     'logistic_guideline_signed'] if part_2_vendor_2 else None
    # sheet['BT105'] = project_dict['vendors'][part_2_vendor_3][
    #     'logistic_guideline_signed'] if part_2_vendor_3 else None
    # sheet['CA105'] = project_dict['vendors'][part_2_vendor_4][
    #     'logistic_guideline_signed'] if part_2_vendor_4 else None
    # sheet['CW105'] = project_dict['vendors'][part_3_vendor_1][
    #     'logistic_guideline_signed'] if part_3_vendor_1 else None
    # sheet['DD105'] = project_dict['vendors'][part_3_vendor_2][
    #     'logistic_guideline_signed'] if part_3_vendor_2 else None
    # sheet['DK105'] = project_dict['vendors'][part_3_vendor_3][
    #     'logistic_guideline_signed'] if part_3_vendor_3 else None
    # sheet['DR105'] = project_dict['vendors'][part_3_vendor_4][
    #     'logistic_guideline_signed'] if part_3_vendor_4 else None
    # sheet['EN105'] = project_dict['vendors'][part_4_vendor_1][
    #     'logistic_guideline_signed'] if part_4_vendor_1 else None
    # sheet['EU105'] = project_dict['vendors'][part_4_vendor_2][
    #     'logistic_guideline_signed'] if part_4_vendor_2 else None
    # sheet['FB105'] = project_dict['vendors'][part_4_vendor_3][
    #     'logistic_guideline_signed'] if part_4_vendor_3 else None
    # sheet['FI105'] = project_dict['vendors'][part_4_vendor_4][
    #     'logistic_guideline_signed'] if part_4_vendor_4 else None

    # # supplier audit result
    # sheet['O110'] = project_dict['vendors'][part_1_vendor_1][
    #     'audit_result'] if part_1_vendor_1 else None
    # sheet['V110'] = project_dict['vendors'][part_1_vendor_2][
    #     'audit_result'] if part_1_vendor_2 else None
    # sheet['AC110'] = project_dict['vendors'][part_1_vendor_3][
    #     'audit_result'] if part_1_vendor_3 else None
    # sheet['AJ110'] = project_dict['vendors'][part_1_vendor_4][
    #     'audit_result'] if part_1_vendor_4 else None
    # sheet['BF110'] = project_dict['vendors'][part_2_vendor_1][
    #     'audit_result'] if part_2_vendor_1 else None
    # sheet['BM110'] = project_dict['vendors'][part_2_vendor_2][
    #     'audit_result'] if part_2_vendor_2 else None
    # sheet['BT110'] = project_dict['vendors'][part_2_vendor_3][
    #     'audit_result'] if part_2_vendor_3 else None
    # sheet['CA110'] = project_dict['vendors'][part_2_vendor_4][
    #     'audit_result'] if part_2_vendor_4 else None
    # sheet['CW110'] = project_dict['vendors'][part_3_vendor_1][
    #     'audit_result'] if part_3_vendor_1 else None
    # sheet['DD110'] = project_dict['vendors'][part_3_vendor_2][
    #     'audit_result'] if part_3_vendor_2 else None
    # sheet['DK110'] = project_dict['vendors'][part_3_vendor_3][
    #     'audit_result'] if part_3_vendor_3 else None
    # sheet['DR110'] = project_dict['vendors'][part_3_vendor_4][
    #     'audit_result'] if part_3_vendor_4 else None
    # sheet['EN110'] = project_dict['vendors'][part_4_vendor_1][
    #     'audit_result'] if part_4_vendor_1 else None
    # sheet['EU110'] = project_dict['vendors'][part_4_vendor_2][
    #     'audit_result'] if part_4_vendor_2 else None
    # sheet['FB110'] = project_dict['vendors'][part_4_vendor_3][
    #     'audit_result'] if part_4_vendor_3 else None
    # sheet['FI110'] = project_dict['vendors'][part_4_vendor_4][
    #     'audit_result'] if part_4_vendor_4 else None

    # # supplier escalation level
    # sheet['O111'] = project_dict['vendors'][part_1_vendor_1][
    #     'escalation_level'] if part_1_vendor_1 else None
    # sheet['V111'] = project_dict['vendors'][part_1_vendor_2][
    #     'escalation_level'] if part_1_vendor_2 else None
    # sheet['AC111'] = project_dict['vendors'][part_1_vendor_3][
    #     'escalation_level'] if part_1_vendor_3 else None
    # sheet['AJ111'] = project_dict['vendors'][part_1_vendor_4][
    #     'escalation_level'] if part_1_vendor_4 else None
    # sheet['BF111'] = project_dict['vendors'][part_2_vendor_1][
    #     'escalation_level'] if part_2_vendor_1 else None
    # sheet['BM111'] = project_dict['vendors'][part_2_vendor_2][
    #     'escalation_level'] if part_2_vendor_2 else None
    # sheet['BT111'] = project_dict['vendors'][part_2_vendor_3][
    #     'escalation_level'] if part_2_vendor_3 else None
    # sheet['CA111'] = project_dict['vendors'][part_2_vendor_4][
    #     'escalation_level'] if part_2_vendor_4 else None
    # sheet['CW111'] = project_dict['vendors'][part_3_vendor_1][
    #     'escalation_level'] if part_3_vendor_1 else None
    # sheet['DD111'] = project_dict['vendors'][part_3_vendor_2][
    #     'escalation_level'] if part_3_vendor_2 else None
    # sheet['DK111'] = project_dict['vendors'][part_3_vendor_3][
    #     'escalation_level'] if part_3_vendor_3 else None
    # sheet['DR111'] = project_dict['vendors'][part_3_vendor_4][
    #     'escalation_level'] if part_3_vendor_4 else None
    # sheet['EN111'] = project_dict['vendors'][part_4_vendor_1][
    #     'escalation_level'] if part_4_vendor_1 else None
    # sheet['EU111'] = project_dict['vendors'][part_4_vendor_2][
    #     'escalation_level'] if part_4_vendor_2 else None
    # sheet['FB111'] = project_dict['vendors'][part_4_vendor_3][
    #     'escalation_level'] if part_4_vendor_3 else None
    # sheet['FI111'] = project_dict['vendors'][part_4_vendor_4][
    #     'escalation_level'] if part_4_vendor_4 else None

    # # supplier ppm_fy
    # sheet['O112'] = project_dict['vendors'][part_1_vendor_1][
    #     'ppm_fy'] if part_1_vendor_1 else None
    # sheet['V112'] = project_dict['vendors'][part_1_vendor_2][
    #     'ppm_fy'] if part_1_vendor_2 else None
    # sheet['AC112'] = project_dict['vendors'][part_1_vendor_3][
    #     'ppm_fy'] if part_1_vendor_3 else None
    # sheet['AJ112'] = project_dict['vendors'][part_1_vendor_4][
    #     'ppm_fy'] if part_1_vendor_4 else None
    # sheet['BF112'] = project_dict['vendors'][part_2_vendor_1][
    #     'ppm_fy'] if part_2_vendor_1 else None
    # sheet['BM112'] = project_dict['vendors'][part_2_vendor_2][
    #     'ppm_fy'] if part_2_vendor_2 else None
    # sheet['BT112'] = project_dict['vendors'][part_2_vendor_3][
    #     'ppm_fy'] if part_2_vendor_3 else None
    # sheet['CA112'] = project_dict['vendors'][part_2_vendor_4][
    #     'ppm_fy'] if part_2_vendor_4 else None
    # sheet['CW112'] = project_dict['vendors'][part_3_vendor_1][
    #     'ppm_fy'] if part_3_vendor_1 else None
    # sheet['DD112'] = project_dict['vendors'][part_3_vendor_2][
    #     'ppm_fy'] if part_3_vendor_2 else None
    # sheet['DK112'] = project_dict['vendors'][part_3_vendor_3][
    #     'ppm_fy'] if part_3_vendor_3 else None
    # sheet['DR112'] = project_dict['vendors'][part_3_vendor_4][
    #     'ppm_fy'] if part_3_vendor_4 else None
    # sheet['EN112'] = project_dict['vendors'][part_4_vendor_1][
    #     'ppm_fy'] if part_4_vendor_1 else None
    # sheet['EU112'] = project_dict['vendors'][part_4_vendor_2][
    #     'ppm_fy'] if part_4_vendor_2 else None
    # sheet['FB112'] = project_dict['vendors'][part_4_vendor_3][
    #     'ppm_fy'] if part_4_vendor_3 else None
    # sheet['FI112'] = project_dict['vendors'][part_4_vendor_4][
    #     'ppm_fy'] if part_4_vendor_4 else None

    # # supplier ppm_target
    # sheet['O113'] = project_dict['vendors'][part_1_vendor_1][
    #     'ppm_target'] if part_1_vendor_1 else None
    # sheet['V113'] = project_dict['vendors'][part_1_vendor_2][
    #     'ppm_target'] if part_1_vendor_2 else None
    # sheet['AC113'] = project_dict['vendors'][part_1_vendor_3][
    #     'ppm_target'] if part_1_vendor_3 else None
    # sheet['AJ113'] = project_dict['vendors'][part_1_vendor_4][
    #     'ppm_target'] if part_1_vendor_4 else None
    # sheet['BF113'] = project_dict['vendors'][part_2_vendor_1][
    #     'ppm_target'] if part_2_vendor_1 else None
    # sheet['BM113'] = project_dict['vendors'][part_2_vendor_2][
    #     'ppm_target'] if part_2_vendor_2 else None
    # sheet['BT113'] = project_dict['vendors'][part_2_vendor_3][
    #     'ppm_target'] if part_2_vendor_3 else None
    # sheet['CA113'] = project_dict['vendors'][part_2_vendor_4][
    #     'ppm_target'] if part_2_vendor_4 else None
    # sheet['CW113'] = project_dict['vendors'][part_3_vendor_1][
    #     'ppm_target'] if part_3_vendor_1 else None
    # sheet['DD113'] = project_dict['vendors'][part_3_vendor_2][
    #     'ppm_target'] if part_3_vendor_2 else None
    # sheet['DK113'] = project_dict['vendors'][part_3_vendor_3][
    #     'ppm_target'] if part_3_vendor_3 else None
    # sheet['DR113'] = project_dict['vendors'][part_3_vendor_4][
    #     'ppm_target'] if part_3_vendor_4 else None
    # sheet['EN113'] = project_dict['vendors'][part_4_vendor_1][
    #     'ppm_target'] if part_4_vendor_1 else None
    # sheet['EU113'] = project_dict['vendors'][part_4_vendor_2][
    #     'ppm_target'] if part_4_vendor_2 else None
    # sheet['FB113'] = project_dict['vendors'][part_4_vendor_3][
    #     'ppm_target'] if part_4_vendor_3 else None
    # sheet['FI113'] = project_dict['vendors'][part_4_vendor_4][
    #     'ppm_target'] if part_4_vendor_4 else None

    # # supplier nqe
    # sheet['O114'] = project_dict['vendors'][part_1_vendor_1][
    #     'nqe'] if part_1_vendor_1 else None
    # sheet['V114'] = project_dict['vendors'][part_1_vendor_2][
    #     'nqe'] if part_1_vendor_2 else None
    # sheet['AC114'] = project_dict['vendors'][part_1_vendor_3][
    #     'nqe'] if part_1_vendor_3 else None
    # sheet['AJ114'] = project_dict['vendors'][part_1_vendor_4][
    #     'nqe'] if part_1_vendor_4 else None
    # sheet['BF114'] = project_dict['vendors'][part_2_vendor_1][
    #     'nqe'] if part_2_vendor_1 else None
    # sheet['BM114'] = project_dict['vendors'][part_2_vendor_2][
    #     'nqe'] if part_2_vendor_2 else None
    # sheet['BT114'] = project_dict['vendors'][part_2_vendor_3][
    #     'nqe'] if part_2_vendor_3 else None
    # sheet['CA114'] = project_dict['vendors'][part_2_vendor_4][
    #     'nqe'] if part_2_vendor_4 else None
    # sheet['CW114'] = project_dict['vendors'][part_3_vendor_1][
    #     'nqe'] if part_3_vendor_1 else None
    # sheet['DD114'] = project_dict['vendors'][part_3_vendor_2][
    #     'nqe'] if part_3_vendor_2 else None
    # sheet['DK114'] = project_dict['vendors'][part_3_vendor_3][
    #     'nqe'] if part_3_vendor_3 else None
    # sheet['DR114'] = project_dict['vendors'][part_3_vendor_4][
    #     'nqe'] if part_3_vendor_4 else None
    # sheet['EN114'] = project_dict['vendors'][part_4_vendor_1][
    #     'nqe'] if part_4_vendor_1 else None
    # sheet['EU114'] = project_dict['vendors'][part_4_vendor_2][
    #     'nqe'] if part_4_vendor_2 else None
    # sheet['FB114'] = project_dict['vendors'][part_4_vendor_3][
    #     'nqe'] if part_4_vendor_3 else None
    # sheet['FI114'] = project_dict['vendors'][part_4_vendor_4][
    #     'nqe'] if part_4_vendor_4 else None

    # # supplier nqe_reimbursed
    # sheet['O115'] = project_dict['vendors'][part_1_vendor_1][
    #     'nqe_reimbursed'] if part_1_vendor_1 else None
    # sheet['V115'] = project_dict['vendors'][part_1_vendor_2][
    #     'nqe_reimbursed'] if part_1_vendor_2 else None
    # sheet['AC115'] = project_dict['vendors'][part_1_vendor_3][
    #     'nqe_reimbursed'] if part_1_vendor_3 else None
    # sheet['AJ115'] = project_dict['vendors'][part_1_vendor_4][
    #     'nqe_reimbursed'] if part_1_vendor_4 else None
    # sheet['BF115'] = project_dict['vendors'][part_2_vendor_1][
    #     'nqe_reimbursed'] if part_2_vendor_1 else None
    # sheet['BM115'] = project_dict['vendors'][part_2_vendor_2][
    #     'nqe_reimbursed'] if part_2_vendor_2 else None
    # sheet['BT115'] = project_dict['vendors'][part_2_vendor_3][
    #     'nqe_reimbursed'] if part_2_vendor_3 else None
    # sheet['CA115'] = project_dict['vendors'][part_2_vendor_4][
    #     'nqe_reimbursed'] if part_2_vendor_4 else None
    # sheet['CW115'] = project_dict['vendors'][part_3_vendor_1][
    #     'nqe_reimbursed'] if part_3_vendor_1 else None
    # sheet['DD115'] = project_dict['vendors'][part_3_vendor_2][
    #     'nqe_reimbursed'] if part_3_vendor_2 else None
    # sheet['DK115'] = project_dict['vendors'][part_3_vendor_3][
    #     'nqe_reimbursed'] if part_3_vendor_3 else None
    # sheet['DR115'] = project_dict['vendors'][part_3_vendor_4][
    #     'nqe_reimbursed'] if part_3_vendor_4 else None
    # sheet['EN115'] = project_dict['vendors'][part_4_vendor_1][
    #     'nqe_reimbursed'] if part_4_vendor_1 else None
    # sheet['EU115'] = project_dict['vendors'][part_4_vendor_2][
    #     'nqe_reimbursed'] if part_4_vendor_2 else None
    # sheet['FB115'] = project_dict['vendors'][part_4_vendor_3][
    #     'nqe_reimbursed'] if part_4_vendor_3 else None
    # sheet['FI115'] = project_dict['vendors'][part_4_vendor_4][
    #     'nqe_reimbursed'] if part_4_vendor_4 else None

    # # supplier safety_representative
    # sheet['O121'] = project_dict['vendors'][part_1_vendor_1][
    #     'safety_representative'] if part_1_vendor_1 else None
    # sheet['V121'] = project_dict['vendors'][part_1_vendor_2][
    #     'safety_representative'] if part_1_vendor_2 else None
    # sheet['AC121'] = project_dict['vendors'][part_1_vendor_3][
    #     'safety_representative'] if part_1_vendor_3 else None
    # sheet['AJ121'] = project_dict['vendors'][part_1_vendor_4][
    #     'safety_representative'] if part_1_vendor_4 else None
    # sheet['BF121'] = project_dict['vendors'][part_2_vendor_1][
    #     'safety_representative'] if part_2_vendor_1 else None
    # sheet['BM121'] = project_dict['vendors'][part_2_vendor_2][
    #     'safety_representative'] if part_2_vendor_2 else None
    # sheet['BT121'] = project_dict['vendors'][part_2_vendor_3][
    #     'safety_representative'] if part_2_vendor_3 else None
    # sheet['CA121'] = project_dict['vendors'][part_2_vendor_4][
    #     'safety_representative'] if part_2_vendor_4 else None
    # sheet['CW121'] = project_dict['vendors'][part_3_vendor_1][
    #     'safety_representative'] if part_3_vendor_1 else None
    # sheet['DD121'] = project_dict['vendors'][part_3_vendor_2][
    #     'safety_representative'] if part_3_vendor_2 else None
    # sheet['DK121'] = project_dict['vendors'][part_3_vendor_3][
    #     'safety_representative'] if part_3_vendor_3 else None
    # sheet['DR121'] = project_dict['vendors'][part_3_vendor_4][
    #     'safety_representative'] if part_3_vendor_4 else None
    # sheet['EN121'] = project_dict['vendors'][part_4_vendor_1][
    #     'safety_representative'] if part_4_vendor_1 else None
    # sheet['EU121'] = project_dict['vendors'][part_4_vendor_2][
    #     'safety_representative'] if part_4_vendor_2 else None
    # sheet['FB121'] = project_dict['vendors'][part_4_vendor_3][
    #     'safety_representative'] if part_4_vendor_3 else None
    # sheet['FI121'] = project_dict['vendors'][part_4_vendor_4][
    #     'safety_representative'] if part_4_vendor_4 else None

    # # decision sourcing approval area
    # # target area
    # sheet['K127'] = project_dict['parts']['part_1']['general_info'][
    #     'nr_id'] or None
    # sheet['AV127'] = project_dict['parts']['part_2']['general_info'][
    #     'nr_id'] or None
    # sheet['CG127'] = project_dict['parts']['part_3']['general_info'][
    #     'nr_id'] or None
    # sheet['DR127'] = project_dict['parts']['part_4']['general_info'][
    #     'nr_id'] or None

    # sheet['K132'] = project_dict['project'][
    #     'customer_nomination_available'] or None
    # sheet['AV132'] = project_dict['project'][
    #     'customer_nomination_available'] or None
    # sheet['CG132'] = project_dict['project'][
    #     'customer_nomination_available'] or None
    # sheet['DR132'] = project_dict['project'][
    #     'customer_nomination_available'] or None

    # sheet['K133'] = project_dict['project']['budget_available'] or None
    # sheet['AV133'] = project_dict['project']['budget_available'] or None
    # sheet['CG133'] = project_dict['project']['budget_available'] or None
    # sheet['DR133'] = project_dict['project']['budget_available'] or None

    # sheet['K135'] = len(project_dict['parts']['part_1']['quotations']) or None
    # sheet['AV135'] = len(project_dict['parts']['part_2']['quotations']) or None
    # sheet['CG135'] = len(project_dict['parts']['part_3']['quotations']) or None
    # sheet['DR135'] = len(project_dict['parts']['part_4']['quotations']) or None

    # # currency
    # sheet['K140'] = project_dict['parts']['part_1']['general_info'][
    #     'currency'] or None
    # sheet['AV140'] = project_dict['parts']['part_2']['general_info'][
    #     'currency'] or None
    # sheet['CG140'] = project_dict['parts']['part_3']['general_info'][
    #     'currency'] or None
    # sheet['DR140'] = project_dict['parts']['part_4']['general_info'][
    #     'currency'] or None

    # # target price
    # sheet['K143'] = project_dict['parts']['part_1']['target_price'][
    #     'target_price100_year_1'] or None
    # sheet['AV143'] = project_dict['parts']['part_2']['target_price'][
    #     'target_price100_year_1'] or None
    # sheet['CG143'] = project_dict['parts']['part_3']['target_price'][
    #     'target_price100_year_1'] or None
    # sheet['DR143'] = project_dict['parts']['part_4']['target_price'][
    #     'target_price100_year_1'] or None

    # # target tool cost
    # sheet['K150'] = project_dict['parts']['part_1']['invest_target'][
    #     'cost_target_tool_1'] or None
    # sheet['AV150'] = project_dict['parts']['part_2']['invest_target'][
    #     'cost_target_tool_1'] or None
    # sheet['CG150'] = project_dict['parts']['part_3']['invest_target'][
    #     'cost_target_tool_1'] or None
    # sheet['DR150'] = project_dict['parts']['part_4']['invest_target'][
    #     'cost_target_tool_1'] or None

    # # target further invest cost
    # sheet['K154'] = project_dict['parts']['part_1']['invest_target'][
    #     'further_invest_cost_tool_1'] or None
    # sheet['AV154'] = project_dict['parts']['part_2']['invest_target'][
    #     'further_invest_cost_tool_1'] or None
    # sheet['CG154'] = project_dict['parts']['part_3']['invest_target'][
    #     'further_invest_cost_tool_1'] or None
    # sheet['DR154'] = project_dict['parts']['part_4']['invest_target'][
    #     'further_invest_cost_tool_1'] or None

    # # target PVO
    # sheet['K159'] = project_dict['parts']['part_1']['general_info'][
    #     'pvo'] or None
    # sheet['AV159'] = project_dict['parts']['part_2']['general_info'][
    #     'pvo'] or None
    # sheet['CG159'] = project_dict['parts']['part_3']['general_info'][
    #     'pvo'] or None
    # sheet['DR159'] = project_dict['parts']['part_4']['general_info'][
    #     'pvo'] or None

    # # quotation area

    # # vendor names
    # sheet['P138'] = project_dict['vendors'][part_1_vendor_1][
    #     'short_name'] if part_1_vendor_1 else None
    # sheet['W138'] = project_dict['vendors'][part_1_vendor_2][
    #     'short_name'] if part_1_vendor_2 else None
    # sheet['AD138'] = project_dict['vendors'][part_1_vendor_3][
    #     'short_name'] if part_1_vendor_3 else None
    # sheet['BA138'] = project_dict['vendors'][part_2_vendor_1][
    #     'short_name'] if part_2_vendor_1 else None
    # sheet['BH138'] = project_dict['vendors'][part_2_vendor_2][
    #     'short_name'] if part_2_vendor_2 else None
    # sheet['BO138'] = project_dict['vendors'][part_2_vendor_3][
    #     'short_name'] if part_2_vendor_3 else None
    # sheet['CL138'] = project_dict['vendors'][part_3_vendor_1][
    #     'short_name'] if part_3_vendor_1 else None
    # sheet['CS138'] = project_dict['vendors'][part_3_vendor_2][
    #     'short_name'] if part_3_vendor_2 else None
    # sheet['CZ138'] = project_dict['vendors'][part_3_vendor_3][
    #     'short_name'] if part_3_vendor_3 else None
    # sheet['DW138'] = project_dict['vendors'][part_4_vendor_1][
    #     'short_name'] if part_4_vendor_1 else None
    # sheet['ED138'] = project_dict['vendors'][part_4_vendor_2][
    #     'short_name'] if part_4_vendor_2 else None
    # sheet['EK138'] = project_dict['vendors'][part_4_vendor_3][
    #     'short_name'] if part_4_vendor_3 else None

    # # quotation currency (use target currency cuz no quotation currency in NR actually)
    # sheet['P139'] = project_dict['parts']['part_1']['general_info'][
    #     'currency'] or None
    # sheet['W139'] = project_dict['parts']['part_1']['general_info'][
    #     'currency'] or None
    # sheet['AD139'] = project_dict['parts']['part_1']['general_info'][
    #     'currency'] or None
    # sheet['BA139'] = project_dict['parts']['part_2']['general_info'][
    #     'currency'] or None
    # sheet['BH139'] = project_dict['parts']['part_2']['general_info'][
    #     'currency'] or None
    # sheet['BO139'] = project_dict['parts']['part_2']['general_info'][
    #     'currency'] or None
    # sheet['CL139'] = project_dict['parts']['part_3']['general_info'][
    #     'currency'] or None
    # sheet['CS139'] = project_dict['parts']['part_3']['general_info'][
    #     'currency'] or None
    # sheet['CZ139'] = project_dict['parts']['part_3']['general_info'][
    #     'currency'] or None
    # sheet['DW139'] = project_dict['parts']['part_4']['general_info'][
    #     'currency'] or None
    # sheet['ED139'] = project_dict['parts']['part_4']['general_info'][
    #     'currency'] or None
    # sheet['EK139'] = project_dict['parts']['part_4']['general_info'][
    #     'currency'] or None

    # # so exchange rate to target quotation should be 1
    # sheet['P141'] = 1 if part_1_vendor_1 else None
    # sheet['W141'] = 1 if part_1_vendor_2 else None
    # sheet['AD141'] = 1 if part_1_vendor_3 else None
    # sheet['BA141'] = 1 if part_2_vendor_1 else None
    # sheet['BH141'] = 1 if part_2_vendor_2 else None
    # sheet['BO141'] = 1 if part_2_vendor_3 else None
    # sheet['CL141'] = 1 if part_3_vendor_1 else None
    # sheet['CS141'] = 1 if part_3_vendor_2 else None
    # sheet['CZ141'] = 1 if part_3_vendor_3 else None
    # sheet['DW141'] = 1 if part_4_vendor_1 else None
    # sheet['ED141'] = 1 if part_4_vendor_2 else None
    # sheet['EK141'] = 1 if part_4_vendor_3 else None

    # # quotations

    # # year 1
    # sheet['P144'] = project_dict['parts']['part_1']['quotations']['vendor_1'][
    #     'prices']['price100_year_1'] or None
    # sheet['W144'] = project_dict['parts']['part_1']['quotations']['vendor_2'][
    #     'prices']['price100_year_1'] or None
    # sheet['AD144'] = project_dict['parts']['part_1']['quotations']['vendor_3'][
    #     'prices']['price100_year_1'] or None
    # sheet['BA144'] = project_dict['parts']['part_2']['quotations']['vendor_1'][
    #     'prices']['price100_year_1'] or None
    # sheet['BH144'] = project_dict['parts']['part_2']['quotations']['vendor_2'][
    #     'prices']['price100_year_1'] or None
    # sheet['BO144'] = project_dict['parts']['part_2']['quotations']['vendor_3'][
    #     'prices']['price100_year_1'] or None
    # sheet['CL144'] = project_dict['parts']['part_3']['quotations']['vendor_1'][
    #     'prices']['price100_year_1'] or None
    # sheet['CS144'] = project_dict['parts']['part_3']['quotations']['vendor_2'][
    #     'prices']['price100_year_1'] or None
    # sheet['CZ144'] = project_dict['parts']['part_3']['quotations']['vendor_3'][
    #     'prices']['price100_year_1'] or None
    # sheet['DW144'] = project_dict['parts']['part_4']['quotations']['vendor_1'][
    #     'prices']['price100_year_1'] or None
    # sheet['ED144'] = project_dict['parts']['part_4']['quotations']['vendor_2'][
    #     'prices']['price100_year_1'] or None
    # sheet['EK144'] = project_dict['parts']['part_4']['quotations']['vendor_3'][
    #     'prices']['price100_year_1'] or None

    # # year2
    # sheet['P145'] = project_dict['parts']['part_1']['quotations']['vendor_1'][
    #     'prices']['price100_year_2'] or None
    # sheet['W145'] = project_dict['parts']['part_1']['quotations']['vendor_2'][
    #     'prices']['price100_year_2'] or None
    # sheet['AD145'] = project_dict['parts']['part_1']['quotations']['vendor_3'][
    #     'prices']['price100_year_2'] or None
    # sheet['BA145'] = project_dict['parts']['part_2']['quotations']['vendor_1'][
    #     'prices']['price100_year_2'] or None
    # sheet['BH145'] = project_dict['parts']['part_2']['quotations']['vendor_2'][
    #     'prices']['price100_year_2'] or None
    # sheet['BO145'] = project_dict['parts']['part_2']['quotations']['vendor_3'][
    #     'prices']['price100_year_2'] or None
    # sheet['CL145'] = project_dict['parts']['part_3']['quotations']['vendor_1'][
    #     'prices']['price100_year_2'] or None
    # sheet['CS145'] = project_dict['parts']['part_3']['quotations']['vendor_2'][
    #     'prices']['price100_year_2'] or None
    # sheet['CZ145'] = project_dict['parts']['part_3']['quotations']['vendor_3'][
    #     'prices']['price100_year_2'] or None
    # sheet['DW145'] = project_dict['parts']['part_4']['quotations']['vendor_1'][
    #     'prices']['price100_year_2'] or None
    # sheet['ED145'] = project_dict['parts']['part_4']['quotations']['vendor_2'][
    #     'prices']['price100_year_2'] or None
    # sheet['EK145'] = project_dict['parts']['part_4']['quotations']['vendor_3'][
    #     'prices']['price100_year_2'] or None

    # # year3
    # sheet['P146'] = project_dict['parts']['part_1']['quotations']['vendor_1'][
    #     'prices']['price100_year_3'] or None
    # sheet['W146'] = project_dict['parts']['part_1']['quotations']['vendor_2'][
    #     'prices']['price100_year_3'] or None
    # sheet['AD146'] = project_dict['parts']['part_1']['quotations']['vendor_3'][
    #     'prices']['price100_year_3'] or None
    # sheet['BA146'] = project_dict['parts']['part_2']['quotations']['vendor_1'][
    #     'prices']['price100_year_3'] or None
    # sheet['BH146'] = project_dict['parts']['part_2']['quotations']['vendor_2'][
    #     'prices']['price100_year_3'] or None
    # sheet['BO146'] = project_dict['parts']['part_2']['quotations']['vendor_3'][
    #     'prices']['price100_year_3'] or None
    # sheet['CL146'] = project_dict['parts']['part_3']['quotations']['vendor_1'][
    #     'prices']['price100_year_3'] or None
    # sheet['CS146'] = project_dict['parts']['part_3']['quotations']['vendor_2'][
    #     'prices']['price100_year_3'] or None
    # sheet['CZ146'] = project_dict['parts']['part_3']['quotations']['vendor_3'][
    #     'prices']['price100_year_3'] or None
    # sheet['DW146'] = project_dict['parts']['part_4']['quotations']['vendor_1'][
    #     'prices']['price100_year_3'] or None
    # sheet['ED146'] = project_dict['parts']['part_4']['quotations']['vendor_2'][
    #     'prices']['price100_year_3'] or None
    # sheet['EK146'] = project_dict['parts']['part_4']['quotations']['vendor_3'][
    #     'prices']['price100_year_3'] or None

    # # year4
    # sheet['P147'] = project_dict['parts']['part_1']['quotations']['vendor_1'][
    #     'prices']['price100_year_4'] or None
    # sheet['W147'] = project_dict['parts']['part_1']['quotations']['vendor_2'][
    #     'prices']['price100_year_4'] or None
    # sheet['AD147'] = project_dict['parts']['part_1']['quotations']['vendor_3'][
    #     'prices']['price100_year_4'] or None
    # sheet['BA147'] = project_dict['parts']['part_2']['quotations']['vendor_1'][
    #     'prices']['price100_year_4'] or None
    # sheet['BH147'] = project_dict['parts']['part_2']['quotations']['vendor_2'][
    #     'prices']['price100_year_4'] or None
    # sheet['BO147'] = project_dict['parts']['part_2']['quotations']['vendor_3'][
    #     'prices']['price100_year_4'] or None
    # sheet['CL147'] = project_dict['parts']['part_3']['quotations']['vendor_1'][
    #     'prices']['price100_year_4'] or None
    # sheet['CS147'] = project_dict['parts']['part_3']['quotations']['vendor_2'][
    #     'prices']['price100_year_4'] or None
    # sheet['CZ147'] = project_dict['parts']['part_3']['quotations']['vendor_3'][
    #     'prices']['price100_year_4'] or None
    # sheet['DW147'] = project_dict['parts']['part_4']['quotations']['vendor_1'][
    #     'prices']['price100_year_4'] or None
    # sheet['ED147'] = project_dict['parts']['part_4']['quotations']['vendor_2'][
    #     'prices']['price100_year_4'] or None
    # sheet['EK147'] = project_dict['parts']['part_4']['quotations']['vendor_3'][
    #     'prices']['price100_year_4'] or None

    # # year5
    # sheet['P148'] = project_dict['parts']['part_1']['quotations']['vendor_1'][
    #     'prices']['price100_year_5'] or None
    # sheet['W148'] = project_dict['parts']['part_1']['quotations']['vendor_2'][
    #     'prices']['price100_year_5'] or None
    # sheet['AD148'] = project_dict['parts']['part_1']['quotations']['vendor_3'][
    #     'prices']['price100_year_5'] or None
    # sheet['BA148'] = project_dict['parts']['part_2']['quotations']['vendor_1'][
    #     'prices']['price100_year_5'] or None
    # sheet['BH148'] = project_dict['parts']['part_2']['quotations']['vendor_2'][
    #     'prices']['price100_year_5'] or None
    # sheet['BO148'] = project_dict['parts']['part_2']['quotations']['vendor_3'][
    #     'prices']['price100_year_5'] or None
    # sheet['CL148'] = project_dict['parts']['part_3']['quotations']['vendor_1'][
    #     'prices']['price100_year_5'] or None
    # sheet['CS148'] = project_dict['parts']['part_3']['quotations']['vendor_2'][
    #     'prices']['price100_year_5'] or None
    # sheet['CZ148'] = project_dict['parts']['part_3']['quotations']['vendor_3'][
    #     'prices']['price100_year_5'] or None
    # sheet['DW148'] = project_dict['parts']['part_4']['quotations']['vendor_1'][
    #     'prices']['price100_year_5'] or None
    # sheet['ED148'] = project_dict['parts']['part_4']['quotations']['vendor_2'][
    #     'prices']['price100_year_5'] or None
    # sheet['EK148'] = project_dict['parts']['part_4']['quotations']['vendor_3'][
    #     'prices']['price100_year_5'] or None

    # # price first tool
    # sheet['P151'] = project_dict['parts']['part_1']['quotations']['vendor_1'][
    #     'invests']['tool_cost_tool_1'] or None
    # sheet['W151'] = project_dict['parts']['part_1']['quotations']['vendor_2'][
    #     'invests']['tool_cost_tool_1'] or None
    # sheet['AD151'] = project_dict['parts']['part_1']['quotations']['vendor_3'][
    #     'invests']['tool_cost_tool_1'] or None
    # sheet['BA151'] = project_dict['parts']['part_2']['quotations']['vendor_1'][
    #     'invests']['tool_cost_tool_1'] or None
    # sheet['BH151'] = project_dict['parts']['part_2']['quotations']['vendor_2'][
    #     'invests']['tool_cost_tool_1'] or None
    # sheet['BO151'] = project_dict['parts']['part_2']['quotations']['vendor_3'][
    #     'invests']['tool_cost_tool_1'] or None
    # sheet['CL151'] = project_dict['parts']['part_3']['quotations']['vendor_1'][
    #     'invests']['tool_cost_tool_1'] or None
    # sheet['CS151'] = project_dict['parts']['part_3']['quotations']['vendor_2'][
    #     'invests']['tool_cost_tool_1'] or None
    # sheet['CZ151'] = project_dict['parts']['part_3']['quotations']['vendor_3'][
    #     'invests']['tool_cost_tool_1'] or None
    # sheet['DW151'] = project_dict['parts']['part_4']['quotations']['vendor_1'][
    #     'invests']['tool_cost_tool_1'] or None
    # sheet['ED151'] = project_dict['parts']['part_4']['quotations']['vendor_2'][
    #     'invests']['tool_cost_tool_1'] or None
    # sheet['EK151'] = project_dict['parts']['part_4']['quotations']['vendor_3'][
    #     'invests']['tool_cost_tool_1'] or None

    # # cavity
    # sheet['P152'] = project_dict['parts']['part_1']['quotations']['vendor_1'][
    #     'invests']['cavity_tool_1'] or None
    # sheet['W152'] = project_dict['parts']['part_1']['quotations']['vendor_2'][
    #     'invests']['cavity_tool_1'] or None
    # sheet['AD152'] = project_dict['parts']['part_1']['quotations']['vendor_3'][
    #     'invests']['cavity_tool_1'] or None
    # sheet['BA152'] = project_dict['parts']['part_2']['quotations']['vendor_1'][
    #     'invests']['cavity_tool_1'] or None
    # sheet['BH152'] = project_dict['parts']['part_2']['quotations']['vendor_2'][
    #     'invests']['cavity_tool_1'] or None
    # sheet['BO152'] = project_dict['parts']['part_2']['quotations']['vendor_3'][
    #     'invests']['cavity_tool_1'] or None
    # sheet['CL152'] = project_dict['parts']['part_3']['quotations']['vendor_1'][
    #     'invests']['cavity_tool_1'] or None
    # sheet['CS152'] = project_dict['parts']['part_3']['quotations']['vendor_2'][
    #     'invests']['cavity_tool_1'] or None
    # sheet['CZ152'] = project_dict['parts']['part_3']['quotations']['vendor_3'][
    #     'invests']['cavity_tool_1'] or None
    # sheet['DW152'] = project_dict['parts']['part_4']['quotations']['vendor_1'][
    #     'invests']['cavity_tool_1'] or None
    # sheet['ED152'] = project_dict['parts']['part_4']['quotations']['vendor_2'][
    #     'invests']['cavity_tool_1'] or None
    # sheet['EK152'] = project_dict['parts']['part_4']['quotations']['vendor_3'][
    #     'invests']['cavity_tool_1'] or None

    # # tool owner, default as Hella
    # sheet['P153'] = 'Hella' if part_1_vendor_1 else None
    # sheet['W153'] = 'Hella' if part_1_vendor_2 else None
    # sheet['AD153'] = 'Hella' if part_1_vendor_3 else None
    # sheet['BA153'] = 'Hella' if part_2_vendor_1 else None
    # sheet['BH153'] = 'Hella' if part_2_vendor_2 else None
    # sheet['BO153'] = 'Hella' if part_2_vendor_3 else None
    # sheet['CL153'] = 'Hella' if part_3_vendor_1 else None
    # sheet['CS153'] = 'Hella' if part_3_vendor_2 else None
    # sheet['CZ153'] = 'Hella' if part_3_vendor_3 else None
    # sheet['DW153'] = 'Hella' if part_4_vendor_1 else None
    # sheet['ED153'] = 'Hella' if part_4_vendor_2 else None
    # sheet['EK153'] = 'Hella' if part_4_vendor_3 else None

    # # price additonal invest
    # sheet['P155'] = project_dict['parts']['part_1']['quotations']['vendor_1'][
    #     'invests']['further_invest_cost_tool_1'] or None
    # sheet['W155'] = project_dict['parts']['part_1']['quotations']['vendor_2'][
    #     'invests']['further_invest_cost_tool_1'] or None
    # sheet['AD155'] = project_dict['parts']['part_1']['quotations']['vendor_3'][
    #     'invests']['further_invest_cost_tool_1'] or None
    # sheet['BA155'] = project_dict['parts']['part_2']['quotations']['vendor_1'][
    #     'invests']['further_invest_cost_tool_1'] or None
    # sheet['BH155'] = project_dict['parts']['part_2']['quotations']['vendor_2'][
    #     'invests']['further_invest_cost_tool_1'] or None
    # sheet['BO155'] = project_dict['parts']['part_2']['quotations']['vendor_3'][
    #     'invests']['further_invest_cost_tool_1'] or None
    # sheet['CL155'] = project_dict['parts']['part_3']['quotations']['vendor_1'][
    #     'invests']['further_invest_cost_tool_1'] or None
    # sheet['CS155'] = project_dict['parts']['part_3']['quotations']['vendor_2'][
    #     'invests']['further_invest_cost_tool_1'] or None
    # sheet['CZ155'] = project_dict['parts']['part_3']['quotations']['vendor_3'][
    #     'invests']['further_invest_cost_tool_1'] or None
    # sheet['DW155'] = project_dict['parts']['part_4']['quotations']['vendor_1'][
    #     'invests']['further_invest_cost_tool_1'] or None
    # sheet['ED155'] = project_dict['parts']['part_4']['quotations']['vendor_2'][
    #     'invests']['further_invest_cost_tool_1'] or None
    # sheet['EK155'] = project_dict['parts']['part_4']['quotations']['vendor_3'][
    #     'invests']['further_invest_cost_tool_1'] or None

    # # price copy tool
    # sheet['P156'] = project_dict['parts']['part_1']['quotations']['vendor_1'][
    #     'invests']['copy_tool_cost_tool_1'] or None
    # sheet['W156'] = project_dict['parts']['part_1']['quotations']['vendor_2'][
    #     'invests']['copy_tool_cost_tool_1'] or None
    # sheet['AD156'] = project_dict['parts']['part_1']['quotations']['vendor_3'][
    #     'invests']['copy_tool_cost_tool_1'] or None
    # sheet['BA156'] = project_dict['parts']['part_2']['quotations']['vendor_1'][
    #     'invests']['copy_tool_cost_tool_1'] or None
    # sheet['BH156'] = project_dict['parts']['part_2']['quotations']['vendor_2'][
    #     'invests']['copy_tool_cost_tool_1'] or None
    # sheet['BO156'] = project_dict['parts']['part_2']['quotations']['vendor_3'][
    #     'invests']['copy_tool_cost_tool_1'] or None
    # sheet['CL156'] = project_dict['parts']['part_3']['quotations']['vendor_1'][
    #     'invests']['copy_tool_cost_tool_1'] or None
    # sheet['CS156'] = project_dict['parts']['part_3']['quotations']['vendor_2'][
    #     'invests']['copy_tool_cost_tool_1'] or None
    # sheet['CZ156'] = project_dict['parts']['part_3']['quotations']['vendor_3'][
    #     'invests']['copy_tool_cost_tool_1'] or None
    # sheet['DW156'] = project_dict['parts']['part_4']['quotations']['vendor_1'][
    #     'invests']['copy_tool_cost_tool_1'] or None
    # sheet['ED156'] = project_dict['parts']['part_4']['quotations']['vendor_2'][
    #     'invests']['copy_tool_cost_tool_1'] or None
    # sheet['EK156'] = project_dict['parts']['part_4']['quotations']['vendor_3'][
    #     'invests']['copy_tool_cost_tool_1'] or None

    # # quotation pvo
    # sheet['P160'] = project_dict['parts']['part_1']['quotations']['vendor_1'][
    #     'pvo'] or None
    # sheet['W160'] = project_dict['parts']['part_1']['quotations']['vendor_2'][
    #     'pvo'] or None
    # sheet['AD160'] = project_dict['parts']['part_1']['quotations']['vendor_3'][
    #     'pvo'] or None
    # sheet['BA160'] = project_dict['parts']['part_2']['quotations']['vendor_1'][
    #     'pvo'] or None
    # sheet['BH160'] = project_dict['parts']['part_2']['quotations']['vendor_2'][
    #     'pvo'] or None
    # sheet['BO160'] = project_dict['parts']['part_2']['quotations']['vendor_3'][
    #     'pvo'] or None
    # sheet['CL160'] = project_dict['parts']['part_3']['quotations']['vendor_1'][
    #     'pvo'] or None
    # sheet['CS160'] = project_dict['parts']['part_3']['quotations']['vendor_2'][
    #     'pvo'] or None
    # sheet['CZ160'] = project_dict['parts']['part_3']['quotations']['vendor_3'][
    #     'pvo'] or None
    # sheet['DW160'] = project_dict['parts']['part_4']['quotations']['vendor_1'][
    #     'pvo'] or None
    # sheet['ED160'] = project_dict['parts']['part_4']['quotations']['vendor_2'][
    #     'pvo'] or None
    # sheet['EK160'] = project_dict['parts']['part_4']['quotations']['vendor_3'][
    #     'pvo'] or None

    # # target PVO - actual PVO
    # sheet['P161'] = (
    #     project_dict['parts']['part_1']['quotations']['vendor_1']['pvo'] -
    #     project_dict['parts']['part_1']['general_info']['pvo']
    # ) if part_1_vendor_1 else None
    # sheet['W161'] = (
    #     project_dict['parts']['part_1']['quotations']['vendor_2']['pvo'] -
    #     project_dict['parts']['part_1']['general_info']['pvo']
    # ) if part_1_vendor_2 else None
    # sheet['AD161'] = (
    #     project_dict['parts']['part_1']['quotations']['vendor_3']['pvo'] -
    #     project_dict['parts']['part_1']['general_info']['pvo']
    # ) if part_1_vendor_3 else None
    # sheet['BA161'] = (
    #     project_dict['parts']['part_2']['quotations']['vendor_1']['pvo'] -
    #     project_dict['parts']['part_2']['general_info']['pvo']
    # ) if part_2_vendor_1 else None
    # sheet['BH161'] = (
    #     project_dict['parts']['part_2']['quotations']['vendor_2']['pvo'] -
    #     project_dict['parts']['part_2']['general_info']['pvo']
    # ) if part_2_vendor_2 else None
    # sheet['BO161'] = (
    #     project_dict['parts']['part_2']['quotations']['vendor_3']['pvo'] -
    #     project_dict['parts']['part_2']['general_info']['pvo']
    # ) if part_2_vendor_3 else None
    # sheet['CL161'] = (
    #     project_dict['parts']['part_3']['quotations']['vendor_1']['pvo'] -
    #     project_dict['parts']['part_3']['general_info']['pvo']
    # ) if part_3_vendor_1 else None
    # sheet['CS161'] = (
    #     project_dict['parts']['part_3']['quotations']['vendor_2']['pvo'] -
    #     project_dict['parts']['part_3']['general_info']['pvo']
    # ) if part_3_vendor_2 else None
    # sheet['CZ161'] = (
    #     project_dict['parts']['part_3']['quotations']['vendor_3']['pvo'] -
    #     project_dict['parts']['part_3']['general_info']['pvo']
    # ) if part_3_vendor_3 else None
    # sheet['DW161'] = (
    #     project_dict['parts']['part_4']['quotations']['vendor_1']['pvo'] -
    #     project_dict['parts']['part_4']['general_info']['pvo']
    # ) if part_4_vendor_1 else None
    # sheet['ED161'] = (
    #     project_dict['parts']['part_4']['quotations']['vendor_2']['pvo'] -
    #     project_dict['parts']['part_4']['general_info']['pvo']
    # ) if part_4_vendor_2 else None
    # sheet['EK161'] = (
    #     project_dict['parts']['part_4']['quotations']['vendor_3']['pvo'] -
    #     project_dict['parts']['part_4']['general_info']['pvo']
    # ) if part_4_vendor_3 else None

    # # capacity check
    # sheet['D188'] = project_dict['project']['sop'][
    #     6:10] if part_1_vendor_1 else None
    # sheet['AY188'] = project_dict['project']['sop'][
    #     6:10] if part_2_vendor_1 else None
    # sheet['CT188'] = project_dict['project']['sop'][
    #     6:10] if part_3_vendor_1 else None
    # sheet['EO188'] = project_dict['project']['sop'][
    #     6:10] if part_4_vendor_1 else None

    # sheet['F188'] = project_dict['project']['year_1_volume'] or None
    # sheet['F189'] = project_dict['project']['year_2_volume'] or None
    # sheet['F190'] = project_dict['project']['year_3_volume'] or None
    # sheet['F191'] = project_dict['project']['year_4_volume'] or None
    # sheet['F192'] = project_dict['project']['year_5_volume'] or None
    # sheet['F193'] = project_dict['project']['year_6_volume'] or None
    # sheet['F194'] = project_dict['project']['year_7_volume'] or None
    # sheet['F195'] = project_dict['project']['year_8_volume'] or None
    # sheet['F196'] = project_dict['project']['year_9_volume'] or None
    # sheet['F197'] = project_dict['project']['year_10_volume'] or None

    # # Team Decision Name
    # sheet['K226'] = project_dict['project']['app'] or None
    # sheet['K227'] = project_dict['project']['pjm'] or None
    # sheet['K228'] = project_dict['parts']['part_1']['general_info'][
    #     'mgs'] or None
    # sheet['K229'] = project_dict['project']['controlling'] or None
    # sheet['K230'] = project_dict['project']['sqa'] or None
    # sheet['K231'] = project_dict['project']['md'] or None
    # sheet['K233'] = project_dict['project']['log'] or None
    # sheet['K234'] = project_dict['project']['me'] or None

    # save the inject
    wb.save('./downloads/' + file_name + '_output.xlsx')
