from pprint import pprint

from openpyxl import load_workbook

import os
from config import TEMPLATE_FOLDER
from app.views import xlsx_inject, docx_inject
from app.models import sql_query, csv_builder, sql_quick_search
from app.models import load_excel, assemble_dict

# test projects
test_project_blank = ""
test_project_none = None
test_project_1 = "1111E.001236"
test_project_2 = "1111P.000099"
test_project_fake = "fake_project_id"

# test single parts

# project_vendor_tuple1
project_vendor_tuple1 = ("1111E.001169", "48200041")
project_vendor_tuple2 = ("1361P.000054", "49100612")
project_vendor_tuple3 = ("1111E.001236", "48201481")

# test project/part tuple
test_project_part_tuple_1 = ("1111E.001236", "178.576-15")
test_project_part_tuple_2 = ("1111E.001236", "191.674-01")
test_project_part_tuple_3 = ("1111E.001236", "229.847-00")
test_project_part_tuple_4 = ("1111E.001236", "fake_part_number")
test_project_part_tuple_5 = ("1111E.001236", "234.536-00")
test_project_part_tuple_6 = ("1111E.001236", "323.140-00")

# test_vendors
test_vendor_1 = ""
test_vendor_2 = "48200025"
test_vendor_3 = "fake_vendor"

# test project part list tuple
test_project_part_list_tuple_1 = ("1111E.001169", ["935.085-00", "935.085-10"])

# test project/vendor/part_list tuple
test_project_vendor_parts_tuple_1 = ("1111E.001169", "48200041",
                                     ["935.085-00", "935.085-10"])

test_project_vendor_parts_tuple_2 = ("1361P.000054", "49100612",
                                     ["187.119-00", "187.120-00", "187.121-00"])

test_project_vendor_parts_tuple_3 = ("1111E.001236", "48201484",
                                     ["234.536-00", "323.140-00"])


test_project_vendor_parts_tuple_4 = ("1111E.001152", "48200053",
                                     ["201.541-00", "201.542-00"])

# test project/part/vendor tuple
test_project_part_vendor_tuple = ("1111E.001169", "935.085-00", "48200041")


# test project info dict

# ALL_PROJECT_LIST = sql.get_all_project_list("1111E.001236", )

def test_show_cwd():
    cwd = os.getcwd()
    print(">>> cwd: ", cwd)


def test_get_project_info():
    assert sql_query.get_project_data_and_info(test_project_1)['project'] == test_project_1
    assert sql_query.get_project_data_and_info(test_project_blank) is None
    assert sql_query.get_project_data_and_info(test_project_fake) is None
    assert sql_query.get_project_data_and_info(test_project_none) is None
    print(sql_query.get_project_data_and_info(test_project_1).keys())
    print("sop date: ", sql_query.get_project_data_and_info(test_project_1)['sop'])


def test_get_part_list_by_project():
    print(sql_query.get_part_list_by_project(test_project_1))
    print(sql_query.get_part_list_by_project(test_project_blank))
    print(sql_query.get_part_list_by_project(test_project_fake))
    print(sql_query.get_part_list_by_project(test_project_none))


def test_get_part_general_info():
    pprint(
        sql_query.get_part_general_info(
            *test_project_part_tuple_5)["plant_short_name"])
    pprint(sql_query.get_part_general_info(*test_project_part_tuple_5)["plant"])
    pprint(sql_query.get_part_general_info(*test_project_part_tuple_5)["mgs"])
    pprint(sql_query.get_part_general_info(*test_project_part_tuple_5)["mgm"])
    pprint(
        sql_query.get_part_general_info(
            *test_project_part_tuple_5)["part"])
    pprint(
        sql_query.get_part_general_info(
            *test_project_part_tuple_6)["part"])


def test_get_part_volume_avg():
    print(">>> volume avg ",
          sql_query.get_part_volume_avg(*test_project_part_tuple_1))
    print(">>> volume avg ",
          sql_query.get_part_volume_avg(*test_project_part_tuple_2))
    print(">>> volume avg ",
          sql_query.get_part_volume_avg(*test_project_part_tuple_3))
    print(">>> volume avg ",
          sql_query.get_part_volume_avg(*test_project_part_tuple_4))


def test_get_part_target_pvo_part():
    print(">>>part_target_pvo_part ",
          sql_query.get_part_target_pvo_part(*test_project_part_tuple_1))
    print(">>>part_target_pvo_part ",
          sql_query.get_part_target_pvo_part(*test_project_part_tuple_2))
    print(">>>part_target_pvo_part ",
          sql_query.get_part_target_pvo_part(*test_project_part_tuple_3))
    print(">>>part_target_pvo_part ",
          sql_query.get_part_target_pvo_part(*test_project_part_tuple_4))


def test_get_part_target_pvo_investment():
    print(">>>part_target_pvo_investment ",
          sql_query.get_part_target_pvo_investment(*test_project_part_tuple_1))
    print(">>>part_target_pvo_investment ",
          sql_query.get_part_target_pvo_investment(*test_project_part_tuple_2))
    print(">>>part_target_pvo_investment ",
          sql_query.get_part_target_pvo_investment(*test_project_part_tuple_3))
    print(">>>part_target_pvo_investment ",
          sql_query.get_part_target_pvo_investment(*test_project_part_tuple_4))


def test_get_part_target_pvo_total():
    print(">>>part_target_pvo_total ",
          sql_query.get_part_target_pvo_total(*test_project_part_tuple_1))
    print(">>>part_target_pvo_total ",
          sql_query.get_part_target_pvo_total(*test_project_part_tuple_2))
    print(">>>part_target_pvo_total ",
          sql_query.get_part_target_pvo_total(*test_project_part_tuple_3))
    print(">>>part_target_pvo_total ",
          sql_query.get_part_target_pvo_total(*test_project_part_tuple_4))


def test_get_part_lifetime():
    print(">>> part lifetime ",
          sql_query.get_part_lifetime(*test_project_part_tuple_1))
    print(">>> part lifetime ",
          sql_query.get_part_lifetime(*test_project_part_tuple_2))
    print(">>> part lifetime ",
          sql_query.get_part_lifetime(*test_project_part_tuple_3))
    print(">>> part lifetime ",
          sql_query.get_part_lifetime(*test_project_part_tuple_4))


def test_get_target_avg_100eur():
    print(">>> target avg price ",
          sql_query.get_part_target_price_avg_100EUR(*test_project_part_tuple_1))
    print(">>> target avg price ",
          sql_query.get_part_target_price_avg_100EUR(*test_project_part_tuple_2))
    print(">>> target avg price ",
          sql_query.get_part_target_price_avg_100EUR(*test_project_part_tuple_3))
    print(">>> target avg price ",
          sql_query.get_part_target_price_avg_100EUR(*test_project_part_tuple_4))


def test_get_part_timing():
    print(">>>part timing ",
          sql_query.get_part_timing(*test_project_part_tuple_2).keys())
    print(">>>part timing ",
          sql_query.get_part_timing(*test_project_part_tuple_2)['ppap_date'])
    ppap = sql_query.get_part_timing(*test_project_part_tuple_2)['ppap_date']
    print("type of ppap: ", type(ppap))


def test_get_vendor_info():
    print(">>>vendor info", sql_query.get_vendor_info(test_vendor_2).keys())
    print(">>>vendor info", sql_query.get_vendor_info(test_vendor_2)['vendor_name'])
    print(">>>vendor info",
          sql_query.get_vendor_info(test_vendor_2)['quality_mgnt_signed'])


def test_get_part_volume_yearly():
    print(">>>volume yearly: ",
          sql_query.get_part_volume_yearly(*test_project_part_tuple_1))
    print(">>>volume yearly: ",
          sql_query.get_part_volume_yearly(*test_project_part_tuple_2))
    print(">>>volume yearly: ",
          sql_query.get_part_volume_yearly(*test_project_part_tuple_3))
    print(">>>volume yearly: ",
          sql_query.get_part_volume_yearly(*test_project_part_tuple_4))


def test_get_part_price_yearly():
    print(">>>price yearly: ",
          sql_query.get_part_price_yearly(*test_project_part_vendor_tuple))


def test_project_sop_eop():
    print(">>>project sop/eop ", sql_query.get_project_sop_eop(test_project_1))
    print(">>>project sop/eop ", sql_query.get_project_sop_eop(test_project_fake))
    print(">>>project sop/eop ", sql_query.get_project_sop_eop(test_project_blank))
    print(">>>project sop/eop ", sql_query.get_project_sop_eop(test_project_none))


# test injection function
def test_xls_inject_risk_eval():
    xlsx_inject.xls_inject_risk_eval(*test_project_part_list_tuple_1)


def test_xls_inject_cbd_single():
    # file prep
    file_name = 'cbd'
    file_path = TEMPLATE_FOLDER + file_name + '.xlsx'

    # load workbook into openpyxl
    wb = load_workbook(file_path)

    # run injection
    xlsx_inject.xls_inject_cbd_single(*test_project_part_tuple_2, "48200041",
                                      wb)


def test_xls_inject_cbd_project():
    xlsx_inject.xls_inject_cbd_project(test_project_1)


def test_xls_inject_ss_single():
    # file prep
    file_name = 'supplier_selection'
    file_path = TEMPLATE_FOLDER + file_name + '.xlsx'

    # load workbook into openpyxl
    wb = load_workbook(file_path)

    # run injection
    xlsx_inject.xls_inject_ss_single(*test_project_part_tuple_2, wb)


def test_xls_inject_ss_project():
    xlsx_inject.xls_inject_ss_project(test_project_1)


# def test_xls_inject_sb():
#     xls_inject.xls_inject_sb(*test_project_part_list_tuple_1)


def test_get_part_volume_weekly():
    print(">>> weekly volume: ",
          sql_query.get_part_volume_weekly(*test_project_part_vendor_tuple))


def test_get_nl_tool_info():
    rows = sql_query.get_nl_tool_info(*test_project_vendor_parts_tuple_2)
    for row in rows:
        print(">>> tool info ")
        print(">>> part info", row['part'], ' ', row['tool'], ' ', row['tool_description'])


def test_get_nl_invest_info():
    rows = sql_query.get_nl_invest_info(*test_project_vendor_parts_tuple_2)
    for row in rows:
        print(">>> invest info ")
        print(">>> part info", row['part'], ' ', row['invest_name'], ' ', row['cost'])


def test_get_project_vendor_qs_yearly():
    print(">>> yearly qs: ", sql_query.get_project_vendor_qs_yearly(*project_vendor_tuple1))
    print(">>> yearly qs: ", sql_query.get_project_vendor_qs_yearly(*project_vendor_tuple2))


def test_assemble_nl_info():
    print(">>> nl info: ",
          assemble_dict.assemble_nl_info(*test_project_vendor_parts_tuple_1))


def test_generate_nl():
    docx_inject.generate_nl(*test_project_vendor_parts_tuple_4)


def test_get_all_project_vendor_tuple():
    print(">>> all project/vendor tuple: ",
          sql_query.get_all_project_vendor_tuple())


def test_all_project_vendor_parts_tuple_for_nl_info():
    for (project, vendor) in sql_query.get_all_project_vendor_tuple():
        part_list = sql_query.get_part_list_by_project_vendor(project, vendor)
        print(">>> nl info: ",
              assemble_dict.assemble_nl_info(project, vendor, part_list))


def test_all_project_vendor_parts_tuple_for_nl_generation():
    for (project, vendor) in sql_query.get_all_project_vendor_tuple():
        part_list = sql_query.get_part_list_by_project_vendor(project, vendor)
        print(">>> project: ", project)
        print(">>> vendor: ", vendor)
        print(">>> part_list: ", part_list)
        docx_inject.generate_nl(project, vendor, part_list)


def test_load_excel():
    load_excel.load_excel("03_project.xlsx")


def test_get_project_info_dict():
    rc = sql_query.project_info_get(test_project_2)
    print(">>> project dict", rc)


def test_create_project_info():
    sql_query.create_project_info_table()


def test_build_csv():
    csv_builder.build_csv("project_info")


def test_generate_ss_folder():
    xlsx_inject.xls_inject_ss_project(test_project_2)


def test_wild_search_project_id():
    rc = sql_quick_search.wild_search_project_by_name(test_project_fake)
    pprint(f"[test_wild_search_project]: name=GM rc={rc}")


def test_wild_search_vendor_id():
    rc = sql_quick_search.wild_search_vendor_by_name("Hongrita")
    pprint(f"[test_wild_search_project]: name=HRT rc={rc}")


def test_search_project_full_info_by_project():
    rc = sql_quick_search.search_project_full_info_by_project(test_project_fake)
    print(f"[search_project_full_info_by_project]: name={test_project_1} rc={rc}")


def test_search_vendor_full_info_by_vendor():
    rc = sql_quick_search.search_vendor_full_info_by_vendor(test_vendor_2)
    print(f"[search_project_full_info_by_project]: name={test_project_2} rc={rc}")

def test_search_part_full_info_by_part_only():
    part = "191.674-01"
    rc = sql_quick_search.search_part_info_by_part(part)
    print(f"[search_project_full_info_by_part]: part={part} rc={rc}")